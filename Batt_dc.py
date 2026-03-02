import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta
import glob
import os

# ─── CONFIG ───
VALID_STOP_CAUSES = ['BATTERY_CAPACITY', 'BATTERY_VOLTAGE', 'MAX_TIME']
CAPACITY_THRESHOLD = 65
VOLT_THRESHOLD = 46.0
ROLLING_WINDOW = 10
RECENT_DAYS = 30
CRITICAL_MIN = 30
DEGRADE_RATIO = 0.7
DURATION_JUMP_RATIO = 1.5

# ─── STEP 1-3: Load & Filter ───
def load_and_filter(filepath):
    df = pd.read_csv(filepath)
    df['start_time'] = pd.to_datetime(df['start_time'])
    df['end_time'] = pd.to_datetime(df['end_time'])
    df = df.sort_values('start_time').reset_index(drop=True)
    mask = (df['start_cause'] == 'SCHEDULED') & (df['stop_cause'].isin(VALID_STOP_CAUSES))
    return df[mask].reset_index(drop=True)

# ─── STEP 4: Detect Battery Replacements ───
def detect_replacements(df):
    """
    Returns list of changepoint indices where battery was likely replaced.
    
    Key principle: A real replacement shows SUSTAINED duration increase over
    multiple discharges, not just a one-off spike.
    
    Method A: Sustained duration jump — compare median of 5 rows before vs 
              5 rows after a candidate point. Both must show >50% increase 
              AND the post-replacement median must be > 60 min.
    Method B: VOLT→CAP stop_cause transition with sustained duration increase.
    
    One-off spikes (e.g., 39→66→43) are NOT replacements.
    Always includes index 0 as the initial segment start.
    """
    changepoints = [0]
    if len(df) < 10:
        return changepoints

    PRE_W = 5   # look-back window
    POST_W = 5  # look-ahead window
    MIN_POST_MEDIAN = 60  # post-replacement must sustain > 60 min
    MIN_RATIO = 1.5       # post/pre median ratio

    durations = df['duration_min'].values

    # Scan for candidate replacement points
    candidates = []
    for i in range(PRE_W, len(df) - POST_W):
        pre_med = np.median(durations[i-PRE_W:i])
        post_med = np.median(durations[i:i+POST_W])

        if pre_med > 0 and post_med / pre_med >= MIN_RATIO and post_med >= MIN_POST_MEDIAN:
            candidates.append((i, post_med / pre_med, post_med))

    # Merge nearby candidates (within 10 rows, keep the best start point)
    if candidates:
        merged = [candidates[0]]
        for c in candidates[1:]:
            if c[0] - merged[-1][0] <= POST_W * 2:
                # Keep the one where the actual jump starts (first in sequence)
                if c[1] > merged[-1][1]:
                    merged[-1] = c
            else:
                merged.append(c)

        for idx, ratio, post_med in merged:
            changepoints.append(idx)

    # Method B: Explicit VOLT→CAP transition with sustained increase
    stops = df['stop_cause'].values
    for i in range(1, len(df) - POST_W):
        if stops[i-1] == 'BATTERY_VOLTAGE' and stops[i] == 'BATTERY_CAPACITY':
            post_med = np.median(durations[i:i+POST_W])
            pre_med = np.median(durations[max(0,i-PRE_W):i])
            if pre_med > 0 and post_med / pre_med >= MIN_RATIO and post_med >= MIN_POST_MEDIAN:
                if i not in changepoints:
                    changepoints.append(i)

    changepoints.sort()
    
    # Final dedup: merge changepoints within 10 rows of each other
    if len(changepoints) > 1:
        deduped = [changepoints[0]]
        for cp in changepoints[1:]:
            if cp - deduped[-1] > POST_W * 2:
                deduped.append(cp)
            # else: skip, too close to previous
        changepoints = deduped
    
    return changepoints

# ─── STEP 5: Detect Degradation ───
def detect_degradation(df_seg):
    """
    Analyze the segment after last replacement.
    Returns: (status, baseline_median, recent_median, details_dict)
    """
    if len(df_seg) < 3:
        return 'INSUFFICIENT_DATA', None, None, {}

    durations = df_seg['duration_min'].values

    # Baseline: rolling window of 10, take the highest median
    baseline = None
    if len(durations) >= ROLLING_WINDOW:
        rolling_medians = []
        for i in range(len(durations) - ROLLING_WINDOW + 1):
            rolling_medians.append(np.median(durations[i:i+ROLLING_WINDOW]))
        baseline = max(rolling_medians)
    else:
        baseline = np.median(durations[:min(10, len(durations))])

    # Recent: median of last 30 days
    last_date = df_seg['start_time'].max()
    cutoff = last_date - timedelta(days=RECENT_DAYS)
    recent = df_seg[df_seg['start_time'] >= cutoff]
    recent_med = recent['duration_min'].median() if len(recent) > 0 else durations[-1]

    # Classify
    if recent_med < CRITICAL_MIN:
        status = 'CRITICAL'
    elif baseline and recent_med < baseline * DEGRADE_RATIO:
        status = 'DEGRADING'
    else:
        status = 'STABLE'

    details = {
        'total_discharges': len(df_seg),
        'first_date': df_seg['start_time'].min(),
        'last_date': df_seg['start_time'].max(),
        'min_duration': durations.min(),
        'max_duration': durations.max(),
        'recent_count': len(recent),
        'pct_drop': round((1 - recent_med / baseline) * 100, 1) if baseline and baseline > 0 else 0
    }

    return status, round(baseline, 1) if baseline else None, round(recent_med, 1), details

# ─── STEP 6: Analyze Strings ───
def analyze_strings(df_seg):
    """
    For each discharge, determine which string 'stopped' the system.
    For BATTERY_CAPACITY: the string that hit 65% first (lowest final_cap_rate)
    For BATTERY_VOLTAGE: the string with lowest final voltage
    Returns per-string analysis.
    """
    # Detect active strings (non-zero init_cap)
    active_strings = []
    for s in [1, 2, 3, 4]:
        col = f'init_cap_rate{s}'
        if col in df_seg.columns and df_seg[col].gt(0).any():
            active_strings.append(s)

    if not active_strings:
        return {}

    string_results = {}
    total = len(df_seg)

    for s in active_strings:
        stopper_count = 0
        for _, row in df_seg.iterrows():
            if row['stop_cause'] == 'BATTERY_CAPACITY':
                # String that hit capacity threshold first = lowest final_cap
                caps = {si: row[f'final_cap_rate{si}'] for si in active_strings}
                min_string = min(caps, key=caps.get)
                if min_string == s:
                    stopper_count += 1
            elif row['stop_cause'] == 'BATTERY_VOLTAGE':
                # String with lowest final voltage
                volts = {si: row[f'final_batt_volt{si}'] for si in active_strings}
                min_string = min(volts, key=volts.get)
                if min_string == s:
                    stopper_count += 1
            elif row['stop_cause'] == 'MAX_TIME':
                pass  # No string is the "stopper"

        stopper_pct = round(stopper_count / total * 100, 1) if total > 0 else 0

        # Classify string health
        # Core rule: If recent duration is critically low (<30 min), 
        # ALL batteries are BAD — even a string with 23% stopper rate
        # is bad because 14 min total runtime means nothing works properly.
        recent_cutoff = df_seg['start_time'].max() - timedelta(days=RECENT_DAYS)
        recent_rows = df_seg[df_seg['start_time'] >= recent_cutoff]
        recent_dur_med = recent_rows['duration_min'].median() if len(recent_rows) > 0 else df_seg['duration_min'].iloc[-5:].median()

        if recent_dur_med < CRITICAL_MIN:
            classification = 'BAD'
        elif stopper_pct > 80:
            classification = 'BAD'
        elif stopper_pct > 50:
            classification = 'WEAK'
        elif stopper_pct > 30:
            classification = 'FAIR'
        else:
            classification = 'GOOD'

        # Average final values for this string
        avg_final_cap = df_seg[f'final_cap_rate{s}'].mean()
        avg_final_volt = df_seg[f'final_batt_volt{s}'].mean()

        string_results[s] = {
            'stopper_count': stopper_count,
            'stopper_pct': stopper_pct,
            'classification': classification,
            'avg_final_cap': round(avg_final_cap, 1),
            'avg_final_volt': round(avg_final_volt, 2),
        }

    return string_results

# ─── STEP 7: Build Excel ───
def build_excel(all_results, output_path):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    # Colors
    fill_header = PatternFill('solid', fgColor='1F4E79')
    fill_critical = PatternFill('solid', fgColor='FF4444')
    fill_degrading = PatternFill('solid', fgColor='FFA500')
    fill_stable = PatternFill('solid', fgColor='00B050')
    fill_bad = PatternFill('solid', fgColor='FF6666')
    fill_weak = PatternFill('solid', fgColor='FFD966')
    fill_fair = PatternFill('solid', fgColor='BDD7EE')
    fill_good = PatternFill('solid', fgColor='C6EFCE')
    fill_light_gray = PatternFill('solid', fgColor='F2F2F2')

    font_header = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    font_bold = Font(bold=True, name='Arial', size=10)
    font_normal = Font(name='Arial', size=10)
    font_status = Font(bold=True, name='Arial', size=10, color='FFFFFF')
    align_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ─── Summary Sheet Headers ───
    headers = [
        'IP', 'Site Location', 'Battery Type',
        'Total Discharges', 'Segment Discharges',
        'Replacements Detected', 'Last Replacement Date',
        'Baseline Duration (min)', 'Recent Duration (min)',
        'Duration Drop %', 'Status',
        'String 1 Status', 'String 1 Stopper%',
        'String 2 Status', 'String 2 Stopper%',
        'First Discharge', 'Last Discharge',
        'Segment Start', 'Segment End'
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    # ─── Write each site result ───
    for row_idx, result in enumerate(all_results, 2):
        values = [
            result['ip'],
            result['site_location'],
            result['batt_type'],
            result['total_discharges'],
            result['seg_discharges'],
            result['num_replacements'],
            result['last_replacement_date'],
            result['baseline'],
            result['recent_med'],
            result['pct_drop'],
            result['status'],
            result.get('str1_class', 'N/A'),
            result.get('str1_pct', ''),
            result.get('str2_class', 'N/A'),
            result.get('str2_pct', ''),
            result['first_date'],
            result['last_date'],
            result['seg_start'],
            result['seg_end'],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = fill_light_gray

        # Color-code status
        status_cell = ws_summary.cell(row=row_idx, column=11)
        if result['status'] == 'CRITICAL':
            status_cell.fill = fill_critical
            status_cell.font = font_status
        elif result['status'] == 'DEGRADING':
            status_cell.fill = fill_degrading
            status_cell.font = Font(bold=True, name='Arial', size=10)
        elif result['status'] == 'STABLE':
            status_cell.fill = fill_stable
            status_cell.font = font_status

        # Color-code string statuses
        for col_offset, key in [(12, 'str1_class'), (14, 'str2_class')]:
            sc = ws_summary.cell(row=row_idx, column=col_offset)
            cls = result.get(key, 'N/A')
            if cls == 'BAD':
                sc.fill = fill_bad
            elif cls == 'WEAK':
                sc.fill = fill_weak
            elif cls == 'FAIR':
                sc.fill = fill_fair
            elif cls == 'GOOD':
                sc.fill = fill_good

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx-1]))
        for row_idx in range(2, len(all_results) + 2):
            val = ws_summary.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 25)

    # Freeze header
    ws_summary.freeze_panes = 'A2'

    # ─── Detail sheet for each site ───
    for result in all_results:
        ip_short = result['ip'].replace('.', '_')
        ws = wb.create_sheet(title=ip_short[:31])

        # Header row
        detail_headers = ['Date', 'Duration (min)', 'Stop Cause',
                         'Final Cap S1', 'Final Cap S2',
                         'Final Volt S1', 'Final Volt S2',
                         'Stopper String', 'Segment']
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        df_seg = result['_df_full']
        cp_last = result['_cp_last']
        active_strings = result['_active_strings']

        for i, (_, row) in enumerate(df_seg.iterrows(), 2):
            date_val = row['start_time'].strftime('%Y-%m-%d')
            ws.cell(row=i, column=1, value=date_val).font = font_normal
            ws.cell(row=i, column=2, value=row['duration_min']).font = font_normal
            ws.cell(row=i, column=3, value=row['stop_cause']).font = font_normal
            ws.cell(row=i, column=4, value=row['final_cap_rate1']).font = font_normal
            ws.cell(row=i, column=5, value=row['final_cap_rate2']).font = font_normal
            ws.cell(row=i, column=6, value=row['final_batt_volt1']).font = font_normal
            ws.cell(row=i, column=7, value=row['final_batt_volt2']).font = font_normal

            # Determine stopper
            stopper = ''
            if row['stop_cause'] == 'BATTERY_CAPACITY' and active_strings:
                caps = {s: row[f'final_cap_rate{s}'] for s in active_strings}
                stopper = f'S{min(caps, key=caps.get)}'
            elif row['stop_cause'] == 'BATTERY_VOLTAGE' and active_strings:
                volts = {s: row[f'final_batt_volt{s}'] for s in active_strings}
                stopper = f'S{min(volts, key=volts.get)}'
            ws.cell(row=i, column=8, value=stopper).font = font_normal

            # Mark segment
            orig_idx = _ 
            seg_label = 'Current' if orig_idx >= cp_last else 'Previous'
            ws.cell(row=i, column=9, value=seg_label).font = font_normal

            # Apply border
            for c in range(1, 10):
                ws.cell(row=i, column=c).border = thin_border
                ws.cell(row=i, column=c).alignment = align_center

            # Highlight low durations
            if row['duration_min'] < CRITICAL_MIN:
                ws.cell(row=i, column=2).fill = fill_critical
                ws.cell(row=i, column=2).font = font_status

        # Auto-width
        for col_idx in range(1, len(detail_headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f'Saved: {output_path}')

# ─── MAIN ───
def process_site(filepath):
    df = load_and_filter(filepath)
    if len(df) == 0:
        return None

    ip = df['IP'].iloc[0]
    site = df['site_location'].iloc[0]
    batt_type = df['batt_type'].iloc[0] if 'batt_type' in df.columns else 'UNKNOWN'

    # Detect replacements
    changepoints = detect_replacements(df)
    num_replacements = len(changepoints) - 1  # subtract initial 0

    # Last segment
    cp_last = changepoints[-1]
    df_seg = df.iloc[cp_last:].reset_index(drop=True)

    last_replacement_date = df.iloc[cp_last]['start_time'].strftime('%Y-%m-%d') if num_replacements > 0 else 'None'

    # Degradation
    status, baseline, recent_med, details = detect_degradation(df_seg)

    # String analysis
    string_results = analyze_strings(df_seg)

    # Active strings
    active_strings = []
    for s in [1, 2, 3, 4]:
        col = f'init_cap_rate{s}'
        if col in df.columns and df[col].gt(0).any():
            active_strings.append(s)

    result = {
        'ip': ip,
        'site_location': site,
        'batt_type': batt_type,
        'total_discharges': len(df),
        'seg_discharges': len(df_seg),
        'num_replacements': num_replacements,
        'last_replacement_date': last_replacement_date,
        'baseline': baseline,
        'recent_med': recent_med,
        'pct_drop': details.get('pct_drop', 0),
        'status': status,
        'first_date': df['start_time'].min().strftime('%Y-%m-%d'),
        'last_date': df['start_time'].max().strftime('%Y-%m-%d'),
        'seg_start': df_seg['start_time'].min().strftime('%Y-%m-%d'),
        'seg_end': df_seg['start_time'].max().strftime('%Y-%m-%d'),
        '_df_full': df,
        '_cp_last': cp_last,
        '_active_strings': active_strings,
    }

    # String details
    for s in [1, 2]:
        if s in string_results:
            result[f'str{s}_class'] = string_results[s]['classification']
            result[f'str{s}_pct'] = string_results[s]['stopper_pct']
        else:
            result[f'str{s}_class'] = 'N/A'
            result[f'str{s}_pct'] = ''

    return result

def main():
    csv_dir = 'data'
    csv_files = sorted(glob.glob(os.path.join(csv_dir, '*.csv')))

    all_results = []
    for f in csv_files:
        print(f'Processing: {os.path.basename(f)}')
        result = process_site(f)
        if result:
            all_results.append(result)
            print(f'  → {result["ip"]} | {result["status"]} | '
                  f'baseline={result["baseline"]}min, recent={result["recent_med"]}min, '
                  f'drop={result["pct_drop"]}%')
            for s in [1, 2]:
                cls = result.get(f'str{s}_class', 'N/A')
                pct = result.get(f'str{s}_pct', '')
                if cls != 'N/A':
                    print(f'  String {s}: {cls} (stopper {pct}%)')

    output_path = 'battery_analysis_report.xlsx'
    build_excel(all_results, output_path)

if __name__ == '__main__':
    main()