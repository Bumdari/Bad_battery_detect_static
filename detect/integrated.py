import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.linear_model import LinearRegression
from datetime import timedelta, datetime
import glob
import os
import warnings
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════
VALID_STOP_CAUSES = ['BATTERY_CAPACITY', 'BATTERY_VOLTAGE', 'MAX_TIME']
ROLLING_WINDOW = 10
RECENT_DAYS = 30
CRITICAL_MIN = 30
DEGRADE_RATIO = 0.7
MEASUREMENT_INTERVAL_DAYS = 2
FORECAST_CYCLES = 15  # 15 цикл = 30 хоног

# ═══════════════════════════════════════════════════════════
# STEP 1-3: Load & Filter
# ═══════════════════════════════════════════════════════════
def load_and_filter(filepath):
    df = pd.read_csv(filepath)
    df['start_time'] = pd.to_datetime(df['start_time'])
    df['end_time'] = pd.to_datetime(df['end_time'])
    df = df.sort_values('start_time').reset_index(drop=True)
    mask = (df['start_cause'] == 'SCHEDULED') & (df['stop_cause'].isin(VALID_STOP_CAUSES))
    return df[mask].reset_index(drop=True)

# ═══════════════════════════════════════════════════════════
# STEP 4: Detect Battery Replacements
# ═══════════════════════════════════════════════════════════
def detect_replacements(df):
    changepoints = [0]
    if len(df) < 10:
        return changepoints

    PRE_W, POST_W = 5, 5
    MIN_POST_MEDIAN, MIN_RATIO = 60, 1.5
    durations = df['duration_min'].values

    candidates = []
    for i in range(PRE_W, len(df) - POST_W):
        pre_med = np.median(durations[i-PRE_W:i])
        post_med = np.median(durations[i:i+POST_W])
        if pre_med > 0 and post_med / pre_med >= MIN_RATIO and post_med >= MIN_POST_MEDIAN:
            candidates.append((i, post_med / pre_med, post_med))

    if candidates:
        merged = [candidates[0]]
        for c in candidates[1:]:
            if c[0] - merged[-1][0] <= POST_W * 2:
                if c[1] > merged[-1][1]:
                    merged[-1] = c
            else:
                merged.append(c)
        for idx, ratio, post_med in merged:
            changepoints.append(idx)

    stops = df['stop_cause'].values
    for i in range(1, len(df) - POST_W):
        if stops[i-1] == 'BATTERY_VOLTAGE' and stops[i] == 'BATTERY_CAPACITY':
            post_med = np.median(durations[i:i+POST_W])
            pre_med = np.median(durations[max(0,i-PRE_W):i])
            if pre_med > 0 and post_med / pre_med >= MIN_RATIO and post_med >= MIN_POST_MEDIAN:
                if i not in changepoints:
                    changepoints.append(i)

    changepoints.sort()
    if len(changepoints) > 1:
        deduped = [changepoints[0]]
        for cp in changepoints[1:]:
            if cp - deduped[-1] > POST_W * 2:
                deduped.append(cp)
        changepoints = deduped
    return changepoints

# ═══════════════════════════════════════════════════════════
# STEP 5: Detect Degradation
# ═══════════════════════════════════════════════════════════
def detect_degradation(df_seg):
    if len(df_seg) < 3:
        return 'INSUFFICIENT_DATA', None, None, {}

    durations = df_seg['duration_min'].values

    if len(durations) >= ROLLING_WINDOW:
        rolling_medians = [np.median(durations[i:i+ROLLING_WINDOW])
                          for i in range(len(durations) - ROLLING_WINDOW + 1)]
        baseline = max(rolling_medians)
    else:
        baseline = np.median(durations[:min(10, len(durations))])

    last_date = df_seg['start_time'].max()
    cutoff = last_date - timedelta(days=RECENT_DAYS)
    recent = df_seg[df_seg['start_time'] >= cutoff]
    recent_med = recent['duration_min'].median() if len(recent) > 0 else durations[-1]

    if recent_med < CRITICAL_MIN:
        status = 'CRITICAL'
    elif baseline and recent_med < baseline * DEGRADE_RATIO:
        status = 'DEGRADING'
    else:
        status = 'STABLE'

    details = {
        'total_discharges': len(df_seg),
        'first_date': df_seg['start_time'].min(),
        'last_date': df_seg['start_time'].max(),
        'min_duration': durations.min(),
        'max_duration': durations.max(),
        'recent_count': len(recent),
        'pct_drop': round((1 - recent_med / baseline) * 100, 1) if baseline and baseline > 0 else 0
    }
    return status, round(baseline, 1) if baseline else None, round(recent_med, 1), details

# ═══════════════════════════════════════════════════════════
# STEP 6: Analyze Strings
# ═══════════════════════════════════════════════════════════
def analyze_strings(df_seg):
    active_strings = []
    for s in [1, 2, 3, 4]:
        col = f'init_cap_rate{s}'
        if col in df_seg.columns and df_seg[col].gt(0).any():
            active_strings.append(s)

    if not active_strings:
        return {}

    string_results = {}
    total = len(df_seg)

    for s in active_strings:
        stopper_count = 0
        for _, row in df_seg.iterrows():
            if row['stop_cause'] == 'BATTERY_CAPACITY':
                caps = {si: row[f'final_cap_rate{si}'] for si in active_strings}
                if min(caps, key=caps.get) == s:
                    stopper_count += 1
            elif row['stop_cause'] == 'BATTERY_VOLTAGE':
                volts = {si: row[f'final_batt_volt{si}'] for si in active_strings}
                if min(volts, key=volts.get) == s:
                    stopper_count += 1

        stopper_pct = round(stopper_count / total * 100, 1) if total > 0 else 0

        recent_cutoff = df_seg['start_time'].max() - timedelta(days=RECENT_DAYS)
        recent_rows = df_seg[df_seg['start_time'] >= recent_cutoff]
        recent_dur_med = recent_rows['duration_min'].median() if len(recent_rows) > 0 else df_seg['duration_min'].iloc[-5:].median()

        if recent_dur_med < CRITICAL_MIN:
            classification = 'BAD'
        elif stopper_pct > 80:
            classification = 'BAD'
        elif stopper_pct > 50:
            classification = 'WEAK'
        elif stopper_pct > 30:
            classification = 'FAIR'
        else:
            classification = 'GOOD'

        avg_final_cap = df_seg[f'final_cap_rate{s}'].mean()
        avg_final_volt = df_seg[f'final_batt_volt{s}'].mean()

        string_results[s] = {
            'stopper_count': stopper_count,
            'stopper_pct': stopper_pct,
            'classification': classification,
            'avg_final_cap': round(avg_final_cap, 1),
            'avg_final_volt': round(avg_final_volt, 2),
        }
    return string_results

# ═══════════════════════════════════════════════════════════
# STEP 7: FORECAST
# ═══════════════════════════════════════════════════════════
def simple_exp_smoothing(values, alpha=0.3):
    result = [values[0]]
    for v in values[1:]:
        result.append(alpha * v + (1 - alpha) * result[-1])
    return result

def forecast_duration(df_seg, n_forecast=FORECAST_CYCLES):
    if len(df_seg) < 5:
        return None

    durations = df_seg['duration_min'].values
    last_date = df_seg['start_time'].max()
    last_cycle = len(df_seg)

    # Арга 1: Local Linear Regression (сүүлийн 15 цикл)
    local_n = min(15, len(df_seg))
    X_local = np.arange(last_cycle - local_n, last_cycle).reshape(-1, 1)
    y_local = durations[-local_n:]

    lr = LinearRegression()
    lr.fit(X_local, y_local)
    slope = lr.coef_[0]

    future_X = np.arange(last_cycle, last_cycle + n_forecast).reshape(-1, 1)
    linear_fc = np.clip(lr.predict(future_X), 0, 200)

    # Арга 2: Exponential Smoothing
    smoothed = simple_exp_smoothing(durations, alpha=0.3)
    last_smooth = smoothed[-1]
    trend_n = min(10, len(smoothed))
    trend_vals = smoothed[-trend_n:]
    trend = (trend_vals[-1] - trend_vals[0]) / (len(trend_vals) - 1) if len(trend_vals) >= 2 else 0
    exp_fc = [max(0, last_smooth + trend * (i + 1)) for i in range(n_forecast)]

    # Ирээдүйн огноонууд
    future_dates = [last_date + timedelta(days=(i + 1) * MEASUREMENT_INTERVAL_DAYS)
                    for i in range(n_forecast)]

    # CRITICAL_MIN-д хүрэх хугацаа
    current_dur = durations[-1]
    cycles_to_critical = None
    critical_date = None

    if slope < 0 and current_dur > CRITICAL_MIN:
        cycles_needed = (CRITICAL_MIN - current_dur) / slope
        if cycles_needed > 0:
            cycles_to_critical = int(cycles_needed)
            critical_date = last_date + timedelta(days=cycles_needed * MEASUREMENT_INTERVAL_DAYS)
    elif current_dur <= CRITICAL_MIN:
        cycles_to_critical = 0
        critical_date = last_date

    return {
        'linear_forecast': linear_fc.tolist(),
        'exp_smooth_forecast': exp_fc,
        'forecast_dates': future_dates,
        'slope': round(slope, 3),
        'current_duration': current_dur,
        'cycles_to_critical': cycles_to_critical,
        'critical_date': critical_date.strftime('%Y-%m-%d') if critical_date else None,
        'last_date': last_date,
    }

# ═══════════════════════════════════════════════════════════
# STEP 8: Build Excel
# ═══════════════════════════════════════════════════════════
def build_excel(all_results, output_path):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    fill_header = PatternFill('solid', fgColor='1F4E79')
    fill_critical = PatternFill('solid', fgColor='FF4444')
    fill_degrading = PatternFill('solid', fgColor='FFA500')
    fill_stable = PatternFill('solid', fgColor='00B050')
    fill_bad = PatternFill('solid', fgColor='FF6666')
    fill_weak = PatternFill('solid', fgColor='FFD966')
    fill_fair = PatternFill('solid', fgColor='BDD7EE')
    fill_good = PatternFill('solid', fgColor='C6EFCE')
    fill_light_gray = PatternFill('solid', fgColor='F2F2F2')
    fill_forecast = PatternFill('solid', fgColor='E8D5F5')

    font_header = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    font_bold = Font(bold=True, name='Arial', size=10)
    font_normal = Font(name='Arial', size=10)
    font_status = Font(bold=True, name='Arial', size=10, color='FFFFFF')
    font_forecast = Font(italic=True, name='Arial', size=10, color='7030A0')
    align_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ─── Summary Sheet ───
    headers = [
        'IP', 'Site Location', 'Battery Type',
        'Total Discharges', 'Segment Discharges',
        'Replacements', 'Last Replacement',
        'Baseline (min)', 'Recent (min)',
        'Drop %', 'Status',
        'Str1 Status', 'Str1 Stopper%',
        'Str2 Status', 'Str2 Stopper%',
        'Trend (min/cycle)', 'Est. Critical Date',
        'Forecast 30d (min)',
        'First Discharge', 'Last Discharge',
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    for row_idx, result in enumerate(all_results, 2):
        fc = result.get('forecast', {})
        values = [
            result['ip'], result['site_location'], result['batt_type'],
            result['total_discharges'], result['seg_discharges'],
            result['num_replacements'], result['last_replacement_date'],
            result['baseline'], result['recent_med'],
            result['pct_drop'], result['status'],
            result.get('str1_class', 'N/A'), result.get('str1_pct', ''),
            result.get('str2_class', 'N/A'), result.get('str2_pct', ''),
            fc.get('slope', ''),
            fc.get('critical_date', 'N/A'),
            round(fc['linear_forecast'][-1], 1) if fc and fc.get('linear_forecast') else '',
            result['first_date'], result['last_date'],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_normal
            cell.alignment = align_center
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = fill_light_gray

        status_cell = ws_summary.cell(row=row_idx, column=11)
        if result['status'] == 'CRITICAL':
            status_cell.fill = fill_critical; status_cell.font = font_status
        elif result['status'] == 'DEGRADING':
            status_cell.fill = fill_degrading; status_cell.font = font_bold
        elif result['status'] == 'STABLE':
            status_cell.fill = fill_stable; status_cell.font = font_status

        fill_map = {'BAD': fill_bad, 'WEAK': fill_weak, 'FAIR': fill_fair, 'GOOD': fill_good}
        for col_offset, key in [(12, 'str1_class'), (14, 'str2_class')]:
            cls = result.get(key, 'N/A')
            if cls in fill_map:
                ws_summary.cell(row=row_idx, column=col_offset).fill = fill_map[cls]

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx-1]))
        for row_idx in range(2, len(all_results) + 2):
            val = ws_summary.cell(row=row_idx, column=col_idx).value
            if val: max_len = max(max_len, len(str(val)))
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 25)
    ws_summary.freeze_panes = 'A2'

    # ─── Detail + Forecast sheets ───
    for result in all_results:
        ip_short = result['ip'].replace('.', '_')
        ws = wb.create_sheet(title=ip_short[:31])

        detail_headers = ['Date', 'Duration (min)', 'Stop Cause',
                         'Final Cap S1', 'Final Cap S2',
                         'Final Volt S1', 'Final Volt S2',
                         'Stopper', 'Segment',
                         'Linear FC', 'ExpSmooth FC']
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = font_header; cell.fill = fill_header
            cell.alignment = align_center; cell.border = thin_border

        df_full = result['_df_full']
        cp_last = result['_cp_last']
        active_strings = result['_active_strings']

        row_num = 2
        for _, row in df_full.iterrows():
            ws.cell(row=row_num, column=1, value=row['start_time'].strftime('%Y-%m-%d')).font = font_normal
            ws.cell(row=row_num, column=2, value=row['duration_min']).font = font_normal
            ws.cell(row=row_num, column=3, value=row['stop_cause']).font = font_normal
            ws.cell(row=row_num, column=4, value=row['final_cap_rate1']).font = font_normal
            ws.cell(row=row_num, column=5, value=row['final_cap_rate2']).font = font_normal
            ws.cell(row=row_num, column=6, value=row['final_batt_volt1']).font = font_normal
            ws.cell(row=row_num, column=7, value=row['final_batt_volt2']).font = font_normal

            stopper = ''
            if row['stop_cause'] == 'BATTERY_CAPACITY' and active_strings:
                caps = {s: row[f'final_cap_rate{s}'] for s in active_strings}
                stopper = f'S{min(caps, key=caps.get)}'
            elif row['stop_cause'] == 'BATTERY_VOLTAGE' and active_strings:
                volts = {s: row[f'final_batt_volt{s}'] for s in active_strings}
                stopper = f'S{min(volts, key=volts.get)}'
            ws.cell(row=row_num, column=8, value=stopper).font = font_normal
            ws.cell(row=row_num, column=9, value='Current' if _ >= cp_last else 'Previous').font = font_normal

            for c in range(1, 12):
                ws.cell(row=row_num, column=c).border = thin_border
                ws.cell(row=row_num, column=c).alignment = align_center

            if row['duration_min'] < CRITICAL_MIN:
                ws.cell(row=row_num, column=2).fill = fill_critical
                ws.cell(row=row_num, column=2).font = font_status
            row_num += 1

        # Forecast rows
        fc = result.get('forecast')
        if fc and fc.get('forecast_dates'):
            ws.cell(row=row_num, column=1, value='── FORECAST ──').font = font_bold
            for c in range(1, 12):
                ws.cell(row=row_num, column=c).fill = PatternFill('solid', fgColor='D9D9D9')
                ws.cell(row=row_num, column=c).border = thin_border
            row_num += 1

            for i, fdate in enumerate(fc['forecast_dates']):
                lin_val = round(fc['linear_forecast'][i], 1)
                exp_val = round(fc['exp_smooth_forecast'][i], 1)
                ws.cell(row=row_num, column=1, value=fdate.strftime('%Y-%m-%d')).font = font_forecast
                ws.cell(row=row_num, column=3, value='FORECAST').font = font_forecast
                ws.cell(row=row_num, column=9, value='Forecast').font = font_forecast
                ws.cell(row=row_num, column=10, value=lin_val).font = font_forecast
                ws.cell(row=row_num, column=11, value=exp_val).font = font_forecast

                for c in range(1, 12):
                    ws.cell(row=row_num, column=c).fill = fill_forecast
                    ws.cell(row=row_num, column=c).border = thin_border
                    ws.cell(row=row_num, column=c).alignment = align_center

                if lin_val < CRITICAL_MIN:
                    ws.cell(row=row_num, column=10).fill = fill_critical
                    ws.cell(row=row_num, column=10).font = font_status
                if exp_val < CRITICAL_MIN:
                    ws.cell(row=row_num, column=11).fill = fill_critical
                    ws.cell(row=row_num, column=11).font = font_status
                row_num += 1

        for col_idx in range(1, len(detail_headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
        ws.freeze_panes = 'A2'

    # ─── Forecast Summary Sheet ───
    ws_fc = wb.create_sheet(title='Forecast Summary')
    fc_headers = ['IP', 'Site', 'Current Duration',
                  'Trend (min/cycle)', 'Trend (min/day)',
                  'Forecast 10d', 'Forecast 20d', 'Forecast 30d',
                  'Status', 'Est. Critical Date', 'Days to Critical']
    for col_idx, h in enumerate(fc_headers, 1):
        cell = ws_fc.cell(row=1, column=col_idx, value=h)
        cell.font = font_header; cell.fill = fill_header
        cell.alignment = align_center; cell.border = thin_border

    for row_idx, result in enumerate(all_results, 2):
        fc = result.get('forecast', {})
        if not fc: continue

        lin_fc = fc.get('linear_forecast', [])
        fc_10d = round(lin_fc[4], 1) if len(lin_fc) > 4 else ''
        fc_20d = round(lin_fc[9], 1) if len(lin_fc) > 9 else ''
        fc_30d = round(lin_fc[-1], 1) if lin_fc else ''
        slope = fc.get('slope', 0)
        slope_per_day = round(slope / MEASUREMENT_INTERVAL_DAYS, 3) if slope else ''

        critical_date = fc.get('critical_date')
        days_to_crit = 'N/A'
        if critical_date and fc.get('last_date'):
            days_to_crit = (datetime.strptime(critical_date, '%Y-%m-%d') - fc['last_date']).days

        values = [result['ip'], result['site_location'],
                  fc.get('current_duration', ''), slope, slope_per_day,
                  fc_10d, fc_20d, fc_30d, result['status'],
                  critical_date or 'N/A', days_to_crit]

        for col_idx, val in enumerate(values, 1):
            cell = ws_fc.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_normal; cell.alignment = align_center; cell.border = thin_border

        st_cell = ws_fc.cell(row=row_idx, column=9)
        if result['status'] == 'CRITICAL':
            st_cell.fill = fill_critical; st_cell.font = font_status
        elif result['status'] == 'DEGRADING':
            st_cell.fill = fill_degrading; st_cell.font = font_bold
        elif result['status'] == 'STABLE':
            st_cell.fill = fill_stable; st_cell.font = font_status

    for col_idx in range(1, len(fc_headers) + 1):
        ws_fc.column_dimensions[get_column_letter(col_idx)].width = 20
    ws_fc.freeze_panes = 'A2'

    wb.save(output_path)
    print(f'Saved: {output_path}')

# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
def process_site(filepath):
    df = load_and_filter(filepath)
    if len(df) == 0:
        return None

    ip = df['IP'].iloc[0]
    site = df['site_location'].iloc[0]
    batt_type = df['batt_type'].iloc[0] if 'batt_type' in df.columns else 'UNKNOWN'

    changepoints = detect_replacements(df)
    num_replacements = len(changepoints) - 1
    cp_last = changepoints[-1]
    df_seg = df.iloc[cp_last:].reset_index(drop=True)
    last_replacement_date = df.iloc[cp_last]['start_time'].strftime('%Y-%m-%d') if num_replacements > 0 else 'None'

    status, baseline, recent_med, details = detect_degradation(df_seg)
    string_results = analyze_strings(df_seg)
    forecast = forecast_duration(df_seg)

    active_strings = [s for s in [1,2,3,4]
                      if f'init_cap_rate{s}' in df.columns and df[f'init_cap_rate{s}'].gt(0).any()]

    result = {
        'ip': ip, 'site_location': site, 'batt_type': batt_type,
        'total_discharges': len(df), 'seg_discharges': len(df_seg),
        'num_replacements': num_replacements, 'last_replacement_date': last_replacement_date,
        'baseline': baseline, 'recent_med': recent_med,
        'pct_drop': details.get('pct_drop', 0), 'status': status,
        'first_date': df['start_time'].min().strftime('%Y-%m-%d'),
        'last_date': df['start_time'].max().strftime('%Y-%m-%d'),
        'seg_start': df_seg['start_time'].min().strftime('%Y-%m-%d'),
        'seg_end': df_seg['start_time'].max().strftime('%Y-%m-%d'),
        'forecast': forecast or {},
        '_df_full': df, '_cp_last': cp_last, '_active_strings': active_strings,
    }

    for s in [1, 2]:
        if s in string_results:
            result[f'str{s}_class'] = string_results[s]['classification']
            result[f'str{s}_pct'] = string_results[s]['stopper_pct']
        else:
            result[f'str{s}_class'] = 'N/A'; result[f'str{s}_pct'] = ''

    return result

def main():
    csv_dir = '../data'
    csv_files = sorted(glob.glob(os.path.join(csv_dir, '*.csv')))

    all_results = []
    for f in csv_files:
        print(f'Processing: {os.path.basename(f)}')
        result = process_site(f)
        if result:
            all_results.append(result)
            fc = result.get('forecast', {})
            print(f'  → {result["ip"]} | {result["status"]} | '
                  f'baseline={result["baseline"]}min, recent={result["recent_med"]}min, '
                  f'drop={result["pct_drop"]}%')
            if fc:
                lin_30 = round(fc["linear_forecast"][-1], 1) if fc.get("linear_forecast") else "?"
                print(f'  Forecast: slope={fc.get("slope")} min/cycle, '
                      f'30d→{lin_30}min, critical={fc.get("critical_date", "N/A")}')
            for s in [1, 2]:
                cls = result.get(f'str{s}_class', 'N/A')
                pct = result.get(f'str{s}_pct', '')
                if cls != 'N/A':
                    print(f'  String {s}: {cls} (stopper {pct}%)')

    output_path = 'battery_analysis_report.xlsx'
    build_excel(all_results, output_path)

if __name__ == '__main__':
    main()