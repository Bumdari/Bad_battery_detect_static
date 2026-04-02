import pandas as pd
import numpy as np
import os
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

data_folder = 'data_archive_2'
OUTPUT_FILE  = 'output_analysis_da2_2.xlsx'

CAP_THRESHOLD   = 65
MIN_DURATION    = 13
BASELINE_TOP_N  = 10
SLOPE_DAYS      = 30
MIN_CYCLES_PRED = 5
MAX_PRED_DAYS   = 365
PRED_WINDOW     = 20

LEVEL_CRITICAL = 30
LEVEL_AVERAGE  = 60
LEVEL_GOOD     = 120
LEVEL_VERYGOOD = 180


def duration_level(dur):
    if dur < LEVEL_CRITICAL:
        return 'CRITICAL'
    elif dur <= LEVEL_AVERAGE:
        return 'AVERAGE'
    elif dur <= LEVEL_GOOD:
        return 'GOOD'
    else:
        return 'VERY GOOD'

def is_valid_cycle(row):
    if row['duration_min'] < MIN_DURATION and row['stop_cause'] != 'BATTERY_VOLTAGE':
        return False
    return True

def weighted_regression(df_window):
    if len(df_window) < MIN_CYCLES_PRED:
        return None, None, None

    x = np.arange(len(df_window), dtype=float)
    y = df_window['duration_min'].values.astype(float)
    w = x + 1

    slope, intercept = np.polyfit(x, y, 1, w=w)

    dates = df_window['start_time'].values
    if len(dates) >= 2:
        total_days = (pd.Timestamp(dates[-1]) - pd.Timestamp(dates[0])).days
        avg_days_per_cycle = total_days / (len(dates) - 1) if total_days > 0 else 2
    else:
        avg_days_per_cycle = 2

    return round(slope, 4), round(intercept, 4), round(avg_days_per_cycle, 2)

def predict_days_to_threshold(current_idx, slope, intercept, threshold, avg_days_per_cycle, today):
    if slope is None or slope >= 0:
        return 'Сайжирч байна'

    current_pred = slope * current_idx + intercept
    if current_pred <= threshold:
        return 'Доошилж эхэлсэн'

    cycles_needed = (current_pred - threshold) / abs(slope)
    days_needed = cycles_needed * avg_days_per_cycle

    if days_needed > MAX_PRED_DAYS:
        return '1 жилд буурахгүй'
    if days_needed <= 0:
        return 'Доошилж эхэлсэн'

    pred_date = today + timedelta(days=int(days_needed))
    return pred_date.strftime('%Y-%m-%d')

def get_trend(slope):
    if slope is None:
        return 'Insufficient data'
    if slope > 0.3:
        return 'IMPROVING'
    elif slope < -0.3:
        return 'DECLINING'
    else:
        return 'STABLE'

def get_group_cap(row, group):
    if group == 1:
        vals = [row['final_cap_rate1'], row['final_cap_rate2']]
    else:
        vals = [row['final_cap_rate3'], row['final_cap_rate4']]
    valid = [v for v in vals if v > 0]
    return round(np.mean(valid), 2) if valid else np.nan

def group_exists(row, group):
    if group == 1:
        return row['init_cap_rate1'] > 0 or row['init_cap_rate2'] > 0
    else:
        return row['init_cap_rate3'] > 0 or row['init_cap_rate4'] > 0

def determine_weak_group(row):
    g1_exists = group_exists(row, 1)
    g2_exists = group_exists(row, 2)
    g1_cap = row['group1_cap']
    g2_cap = row['group2_cap']
    cause = row['stop_cause']

    if g1_exists and not g2_exists:
        return 'single_group1'
    if g2_exists and not g1_exists:
        return 'single_group2'

    if cause in ['BATTERY_VOLTAGE', 'MAX_TIME']:
        if pd.isna(g1_cap) or pd.isna(g2_cap):
            return 'unknown'
        if g1_cap < g2_cap:
            return 'group1'
        elif g2_cap < g1_cap:
            return 'group2'
        else:
            return 'equal'
    elif cause == 'BATTERY_CAPACITY':
        g1_weak = (not pd.isna(g1_cap)) and (g1_cap <= CAP_THRESHOLD)
        g2_weak = (not pd.isna(g2_cap)) and (g2_cap <= CAP_THRESHOLD)
        if g1_weak and not g2_weak:
            return 'group1'
        elif g2_weak and not g1_weak:
            return 'group2'
        elif g1_weak and g2_weak:
            return 'equal'
        else:
            if g1_cap < g2_cap:
                return 'group1'
            elif g2_cap < g1_cap:
                return 'group2'
            else:
                return 'equal'
    return 'unknown'


all_files = [f for f in os.listdir(data_folder) if f.lower().endswith('.csv')]
print(f"{len(all_files)} ширхэг csv")

results = []
today = pd.Timestamp.today().normalize()

for filename in all_files:
    filepath = os.path.join(data_folder, filename)
    try:
        df = pd.read_csv(filepath)
        df['start_time'] = pd.to_datetime(df['start_time'])
        df = df.sort_values('start_time').reset_index(drop=True)

        valid_causes = ['BATTERY_VOLTAGE', 'BATTERY_CAPACITY', 'MAX_TIME']
        df = df[df['stop_cause'].isin(valid_causes)].reset_index(drop=True)
        if len(df) < 1:
            print(f" → {filename}: хоосон, алгасав")
            continue

        df = df[df.apply(is_valid_cycle, axis=1)].reset_index(drop=True)
        if len(df) < 1:
            print(f" → {filename}: шүүлтийн дараа хоосон")
            continue

        df['group1_cap'] = df.apply(lambda r: get_group_cap(r, 1), axis=1)
        df['group2_cap'] = df.apply(lambda r: get_group_cap(r, 2), axis=1)
        df['weak_group'] = df.apply(determine_weak_group, axis=1)

        battery_change_indices = [0]
        durations = df['duration_min']
        for i in range(1, len(df)):
            if i >= 5 and i + 5 <= len(df):
                pre_med = durations.iloc[i-5:i].median()
                post_med = durations.iloc[i:i+5].median()
                if pre_med > 0 and post_med / pre_med >= 1.5 and post_med >= 60:
                    battery_change_indices.append(i)

        last_change_idx = battery_change_indices[-1]
        battery_change_date = df.loc[last_change_idx, 'start_time']
        recent_df = df.iloc[last_change_idx:].copy().reset_index(drop=True)

        top_10 = recent_df['duration_min'].sort_values(ascending=False).head(BASELINE_TOP_N)
        baseline = top_10.median()

        cutoff_date = df['start_time'].max() - pd.Timedelta(days=SLOPE_DAYS)
        last_30_df = df[df['start_time'] >= cutoff_date]
        slope_30 = last_30_df['duration_min'].median() if len(last_30_df) > 0 else recent_df['duration_min'].median()

        last_row = df.iloc[-1]
        last_duration   = last_row['duration_min']
        last_stop_cause = last_row['stop_cause']
        last_start_date = last_row['start_time'].strftime('%Y-%m-%d')
        ip              = df['IP'].iloc[0]

        drop_percent = round(((baseline - slope_30) / baseline * 100), 2) if baseline > 0 else 0.0
        if drop_percent < 25:
            drop_level = 'NORMAL'
        elif drop_percent <= 55:
            drop_level = 'STRONG'
        else:
            drop_level = 'CRITICAL'

        if last_duration >= 59:
            status = 'stable'
        elif last_duration < 30:
            status = 'critical'
        else:
            status = 'degrading'

        min_idx      = recent_df['duration_min'].idxmin()
        min_duration = recent_df.loc[min_idx, 'duration_min']
        min_stop     = recent_df.loc[min_idx, 'stop_cause']
        min_date     = recent_df.loc[min_idx, 'start_time'].strftime('%Y-%m-%d')

        current_level = duration_level(last_duration)

        pred_window_df = recent_df.tail(PRED_WINDOW).reset_index(drop=True)
        wlr_slope, wlr_intercept, avg_days_cycle = weighted_regression(pred_window_df)
        trend = get_trend(wlr_slope)
        current_idx = len(pred_window_df)

        has_enough = wlr_slope is not None and len(pred_window_df) >= MIN_CYCLES_PRED

        if current_level == 'VERY GOOD' and has_enough:
            pred_to_good     = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_GOOD,     avg_days_cycle, today)
            pred_to_average  = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_AVERAGE,  avg_days_cycle, today)
            pred_to_critical = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_CRITICAL, avg_days_cycle, today)
        elif current_level == 'GOOD' and has_enough:
            pred_to_good     = 'Одоогийн түвшин'
            pred_to_average  = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_AVERAGE,  avg_days_cycle, today)
            pred_to_critical = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_CRITICAL, avg_days_cycle, today)
        elif current_level == 'AVERAGE' and has_enough:
            pred_to_good     = 'Муу'
            pred_to_average  = 'Одоогийн түвшин'
            pred_to_critical = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_CRITICAL, avg_days_cycle, today)
        elif current_level == 'CRITICAL':
            pred_to_good     = 'Хурцадмал'
            pred_to_average  = 'Хурцадмал'
            pred_to_critical = 'Хурцадмал'
        else:
            pred_to_good     = 'Өгөгдөл хүрэлцэхгүй'
            pred_to_average  = 'Өгөгдөл хүрэлцэхгүй'
            pred_to_critical = 'Өгөгдөл хүрэлцэхгүй'

        NEAR_CRITICAL_WINDOW = 10
        NEAR_CRITICAL_COUNT  = 2
        recent_10 = recent_df.tail(NEAR_CRITICAL_WINDOW)
        critical_hits = (recent_10['duration_min'] < LEVEL_CRITICAL).sum()
        near_critical = critical_hits >= NEAR_CRITICAL_COUNT

        if near_critical:
            last_pred = recent_10['duration_min'].mean()
            if wlr_slope is not None and wlr_slope < 0:
                cycles_needed = (last_pred - LEVEL_CRITICAL) / abs(wlr_slope)
                days_needed = cycles_needed * avg_days_cycle
                if 0 < days_needed <= MAX_PRED_DAYS:
                    near_critical_date = (today + timedelta(days=int(days_needed))).strftime('%Y-%m-%d')
                else:
                    near_critical_date = 'Тодорхойгүй'
            else:
                near_critical_date = (today + timedelta(days=int(avg_days_cycle * 3))).strftime('%Y-%m-%d')
        else:
            near_critical_date = None

        trend_arrow = '↘' if trend == 'DECLINING' else ('↗' if trend == 'IMPROVING' else '→')

        if current_level == 'VERY GOOD':
            next_level = 'GOOD'
            next_date  = pred_to_good
        elif current_level == 'GOOD':
            next_level = 'AVERAGE'
            next_date  = pred_to_average
        elif current_level == 'AVERAGE':
            next_level = 'CRITICAL'
            next_date  = pred_to_critical
        else:
            next_level = None
            next_date  = None

        if current_level == 'CRITICAL':
            forecast_str = 'CRITICAL ↘ Хурцадмал'
        elif near_critical:
            forecast_str = f"{current_level} ↘ Аль хэдийн CRITICAL орчимд байна ({near_critical_date})"
        elif next_level and next_date and next_date not in (
                'Сайжирч байна', 'Доошилж эхэлсэн', '1 жилд буурахгүй',
                'Муу', 'Одоогийн түвшин', 'Хурцадмал', 'Өгөгдөл хүрэлцэхгүй'):
            forecast_str = f"{current_level} {trend_arrow} {next_level} ({next_date})"
        elif trend in ('IMPROVING', 'STABLE'):
            forecast_str = f"{current_level} {trend_arrow} Хэвийн"
        else:
            forecast_str = f"{current_level} {trend_arrow} {next_date or '—'}"

        last_30_weak  = df[df['start_time'] >= cutoff_date]['weak_group'].value_counts()
        dominant_weak = last_30_weak.index[0] if len(last_30_weak) > 0 else 'unknown'

        results.append({
            'ip address'               : ip,
            'battery change date'      : battery_change_date.strftime('%Y-%m-%d'),
            'baseline'                 : round(baseline, 2),
            'slope(30d median)'        : round(slope_30, 2),
            'drop percent'             : drop_percent,
            'last start date'          : last_start_date,
            'last duration_min'        : last_duration,
            'last stop cause'          : last_stop_cause,
            'critical degrading stable': status,
            'status & forecast'        : forecast_str,
            'min duration'             : min_duration,
            'min duration date'        : min_date,
            'min stop cause'           : min_stop,
            'last weak group'          : last_row['weak_group'],
            'dominant weak 30d'        : dominant_weak,
            '_drop_level'              : drop_level,
        })

        print(f" ✓ {filename} → {ip} | {current_level} | {trend} | drop: {drop_percent}%")

    except Exception as e:
        print(f" ✗ {filename} алдаа: {e}")

if not results:
    print("Ямар ч файл боловсруулж чадсангүй!")
else:
    output_df = pd.DataFrame(results)

    status_order = {'critical': 0, 'degrading': 1, 'stable': 2}
    output_df['_sort'] = output_df['critical degrading stable'].map(status_order)
    output_df = output_df.sort_values('_sort').drop(columns='_sort').reset_index(drop=True)

    drop_level_map = output_df['_drop_level'].tolist()
    output_df = output_df.drop(columns=['_drop_level'])

    output_df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    col_map = {}
    for c in range(1, ws.max_column + 1):
        col_map[ws.cell(1, c).value] = c

    red    = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
    orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    yellow = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    green  = PatternFill(start_color='00CC66', end_color='00CC66', fill_type='solid')
    gray   = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    for i, row in enumerate(range(2, ws.max_row + 1)):

        if 'critical degrading stable' in col_map:
            val = ws.cell(row, col_map['critical degrading stable']).value
            if val == 'critical':    ws.cell(row, col_map['critical degrading stable']).fill = red
            elif val == 'degrading': ws.cell(row, col_map['critical degrading stable']).fill = orange
            elif val == 'stable':    ws.cell(row, col_map['critical degrading stable']).fill = green

        if 'drop percent' in col_map:
            dl = drop_level_map[i]
            if dl == 'CRITICAL': ws.cell(row, col_map['drop percent']).fill = red
            elif dl == 'STRONG': ws.cell(row, col_map['drop percent']).fill = orange
            else:                ws.cell(row, col_map['drop percent']).fill = green

        if 'status & forecast' in col_map:
            val = ws.cell(row, col_map['status & forecast']).value or ''
            if val.startswith('CRITICAL') or 'CRITICAL орчимд' in val:
                ws.cell(row, col_map['status & forecast']).fill = red
            elif val.startswith('AVERAGE'):   ws.cell(row, col_map['status & forecast']).fill = orange
            elif val.startswith('GOOD'):      ws.cell(row, col_map['status & forecast']).fill = yellow
            elif val.startswith('VERY GOOD'): ws.cell(row, col_map['status & forecast']).fill = green

        for col_name in ['last weak group', 'dominant weak 30d']:
            if col_name in col_map:
                val = ws.cell(row, col_map[col_name]).value
                if val == 'group1':                    ws.cell(row, col_map[col_name]).fill = orange
                elif val == 'group2':                  ws.cell(row, col_map[col_name]).fill = yellow
                elif val == 'equal':                   ws.cell(row, col_map[col_name]).fill = red
                elif val and val.startswith('single'): ws.cell(row, col_map[col_name]).fill = gray

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = header_fill
        ws.cell(1, c).font = header_font
        ws.cell(1, c).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 24

    ws.row_dimensions[1].height = 35
    wb.save(OUTPUT_FILE)

    print(f"\n{'='*55}")
    print(f"ДУУСЛАА: {len(results)} IP → {OUTPUT_FILE}")

    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email import encoders
    from datetime import datetime

    SMTP_HOST = "10.36.66.46"
    SMTP_PORT = 25
    FROM_EMAIL = "bumdari.b@mobicom.mn"            
    TO_EMAILS  = ["tsogtbaatar.e@mobicom.mn"]      
    CC_EMAILS  = []                            

    today_str = datetime.today().strftime('%Y-%m-%d')
    subject   = f"Баттерийн анализ тайлан — {today_str}"
    body      = f"""Сайн байна уу,

{today_str}-ны өдрийн UPS баттерийн анализ тайланг хавсаргав.

• Баттерийн ерөнхий байдал (critical / degrading / stable)
• Drop percent болон drop level
• Status & Forecast — цаашдын таамаглал

Хүндэтгэлтэй,
Автомат мэдэгдэл"""

    msg = MIMEMultipart()
    msg['From']    = FROM_EMAIL
    msg['To']      = ', '.join(TO_EMAILS)
    if CC_EMAILS:
        msg['CC']  = ', '.join(CC_EMAILS)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))

    with open(OUTPUT_FILE, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="battery_report_{today_str}.xlsx"')
        msg.attach(part)

    all_recipients = TO_EMAILS + CC_EMAILS
    try:
        server = smtplib.SMTP()
        server.set_debuglevel(1)
        code, msg_b = server.connect(SMTP_HOST, SMTP_PORT)
        print(f"Connected: {code} {msg_b}")
        server.ehlo()
        print(f"EHLO response: {server.ehlo_resp}")
        server.sendmail(FROM_EMAIL, all_recipients, msg.as_string())
        server.quit()
        print(f"✓ Имэйл амжилттай илгээгдлээ → {', '.join(TO_EMAILS)}")
    except smtplib.SMTPConnectError as e:
        print(f"✗ Холболт амжилтгүй: {e}")
    except smtplib.SMTPException as e:
        print(f"✗ SMTP алдаа: {e}")
    except Exception as e:
        print(f"✗ Алдаа: {type(e).__name__}: {e}")
    except Exception as e:
        print(f"✗ Имэйл илгээхэд алдаа гарлаа: {e}")