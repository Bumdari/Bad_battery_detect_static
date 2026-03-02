"""
Battery Fleet Report Generator
================================
Processes all site CSV files from a data folder and generates
a single Excel report with:
  - Summary dashboard (all sites at a glance)
  - Per-site detail sheets
  - Color-coded status indicators
  - Forecasts and threshold estimates
"""

import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os, glob, warnings, logging

warnings.filterwarnings('ignore')

log = logging.getLogger('battery_report')
if not log.handlers:
    h = logging.StreamHandler()
    h.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S'))
    log.addHandler(h)
log.setLevel(logging.INFO)

# ── Config ──
MEASUREMENT_INTERVAL_DAYS = 2
SOH_WARNING = 80.0
SOH_CRITICAL = 60.0
REPLACEMENT_JUMP = 15.0
REPLACEMENT_LOOKBACK = 3
REPLACEMENT_COOLDOWN = 5
FORECAST_HORIZON = 10
LOCAL_REGRESSION_WINDOW = 8

# ── Styles ──
FONT_HEADER = Font(name='Arial', bold=True, color='FFFFFF', size=11)
FONT_NORMAL = Font(name='Arial', size=10)
FONT_BOLD = Font(name='Arial', bold=True, size=10)
FONT_TITLE = Font(name='Arial', bold=True, size=14, color='1F4E79')
FONT_SUBTITLE = Font(name='Arial', bold=True, size=11, color='1F4E79')

FILL_HEADER = PatternFill('solid', fgColor='1F4E79')
FILL_OK = PatternFill('solid', fgColor='C6EFCE')
FILL_WATCH = PatternFill('solid', fgColor='FFEB9C')
FILL_WARNING = PatternFill('solid', fgColor='FFC7CE')
FILL_CRITICAL = PatternFill('solid', fgColor='FF4444')
FILL_LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
FILL_WHITE = PatternFill('solid', fgColor='FFFFFF')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


# ═══════════════════════════════════════════════════════════
# Analysis Functions
# ═══════════════════════════════════════════════════════════

def detect_replacements(soh_values, jump_threshold=REPLACEMENT_JUMP, lookback=REPLACEMENT_LOOKBACK):
    if len(soh_values) < lookback + 1:
        return []
    replacements = []
    cooldown = lookback + 2
    for i in range(lookback, len(soh_values)):
        if replacements and (i - replacements[-1]) < cooldown:
            continue
        lb_start = max(0, i - lookback)
        recent_min = np.nanmin(soh_values[lb_start:i])
        cur = soh_values[i]
        if np.isnan(cur) or np.isnan(recent_min):
            continue
        if cur - recent_min >= jump_threshold:
            replacements.append(i)
    return replacements


def analyze_single_battery(df, soh_col, battery_idx):
    """Analyze one battery string. Returns dict with all metrics."""
    soh_raw = df[soh_col].copy()
    soh_raw = soh_raw.replace(0, np.nan)
    soh_raw = soh_raw.clip(lower=0, upper=100)

    # IQR outlier removal
    valid = soh_raw.dropna()
    if len(valid) > 5:
        q1, q3 = valid.quantile(0.25), valid.quantile(0.75)
        iqr = q3 - q1
        soh_raw = soh_raw.where((soh_raw >= q1 - 2*iqr) & (soh_raw <= q3 + 2*iqr))

    soh = soh_raw.interpolate(method='linear', limit=3, limit_direction='both').ffill().bfill()

    if soh.isna().all() or len(soh.dropna()) < 3:
        return None

    # Replacement detection
    replacements = detect_replacements(soh.values)
    n_replacements = len(replacements)

    # Current segment (after last replacement)
    seg_start = replacements[-1] if replacements else 0
    seg_soh = soh.iloc[seg_start:].values
    seg_dates = df['start_time'].iloc[seg_start:]
    seg_len = len(seg_soh)

    current_soh = float(seg_soh[-1])
    max_soh = float(np.nanmax(seg_soh))
    min_soh = float(np.nanmin(seg_soh))
    mean_soh = float(np.nanmean(seg_soh))

    # Degradation rate
    rate_per_cycle = 0.0
    rate_per_day = 0.0
    r_squared = 0.0

    if seg_len >= 3:
        recent_n = min(LOCAL_REGRESSION_WINDOW, seg_len)
        X = np.arange(recent_n).reshape(-1, 1)
        y = seg_soh[-recent_n:]
        lr = LinearRegression().fit(X, y)
        rate_per_cycle = float(lr.coef_[0])
        rate_per_day = rate_per_cycle / MEASUREMENT_INTERVAL_DAYS
        ss_res = np.sum((y - lr.predict(X)) ** 2)
        ss_tot = np.sum((y - y.mean()) ** 2)
        r_squared = float(1 - ss_res / ss_tot) if ss_tot > 0 else 0.0

    # Degradation acceleration
    accelerating = False
    if seg_len >= 20:
        h = seg_len // 2
        X1 = np.arange(h).reshape(-1, 1)
        X2 = np.arange(seg_len - h).reshape(-1, 1)
        r1 = LinearRegression().fit(X1, seg_soh[:h]).coef_[0]
        r2 = LinearRegression().fit(X2, seg_soh[h:]).coef_[0]
        accelerating = r2 < r1

    # Time to thresholds
    cycles_to_80 = None
    days_to_80 = None
    date_to_80 = None
    cycles_to_60 = None
    days_to_60 = None
    date_to_60 = None

    last_date = df['start_time'].iloc[-1]

    if rate_per_cycle < 0 and current_soh > SOH_WARNING:
        c = (SOH_WARNING - current_soh) / rate_per_cycle
        if c > 0:
            cycles_to_80 = c
            days_to_80 = c * MEASUREMENT_INTERVAL_DAYS
            date_to_80 = last_date + timedelta(days=days_to_80)

    if rate_per_cycle < 0 and current_soh > SOH_CRITICAL:
        c = (SOH_CRITICAL - current_soh) / rate_per_cycle
        if c > 0:
            cycles_to_60 = c
            days_to_60 = c * MEASUREMENT_INTERVAL_DAYS
            date_to_60 = last_date + timedelta(days=days_to_60)

    # Status
    if current_soh < SOH_CRITICAL:
        status = 'CRITICAL'
    elif current_soh < SOH_WARNING:
        status = 'WARNING'
    elif cycles_to_80 is not None and cycles_to_80 < 30:
        status = 'WATCH'
    else:
        status = 'OK'

    # Forecast
    forecast_soh = []
    forecast_dates = []
    if seg_len >= 5:
        n_fc = min(LOCAL_REGRESSION_WINDOW, seg_len)
        X_fc = np.arange(n_fc).reshape(-1, 1)
        y_fc = seg_soh[-n_fc:]
        lr_fc = LinearRegression().fit(X_fc, y_fc)
        for i in range(1, FORECAST_HORIZON + 1):
            pred = float(lr_fc.predict(np.array([[n_fc - 1 + i]]))[0])
            forecast_soh.append(np.clip(pred, 0, 100))
            forecast_dates.append(last_date + timedelta(days=i * MEASUREMENT_INTERVAL_DAYS))

    # Voltage drop info
    init_v_col = f'init_batt_volt{battery_idx}'
    final_v_col = f'final_batt_volt{battery_idx}'
    avg_volt_drop = None
    if init_v_col in df.columns and final_v_col in df.columns:
        vd = (df[init_v_col] - df[final_v_col]).clip(lower=0, upper=10)
        avg_volt_drop = float(vd.iloc[seg_start:].mean())

    # Replacement dates
    replacement_dates = [df['start_time'].iloc[r] for r in replacements if r < len(df)]

    return {
        'current_soh': current_soh,
        'max_soh': max_soh,
        'min_soh': min_soh,
        'mean_soh': mean_soh,
        'rate_per_cycle': rate_per_cycle,
        'rate_per_day': rate_per_day,
        'r_squared': r_squared,
        'accelerating': accelerating,
        'n_replacements': n_replacements,
        'replacement_dates': replacement_dates,
        'current_segment_cycles': seg_len,
        'total_cycles': len(soh),
        'cycles_to_80': cycles_to_80,
        'days_to_80': days_to_80,
        'date_to_80': date_to_80,
        'cycles_to_60': cycles_to_60,
        'days_to_60': days_to_60,
        'date_to_60': date_to_60,
        'status': status,
        'forecast_soh': forecast_soh,
        'forecast_dates': forecast_dates,
        'avg_volt_drop': avg_volt_drop,
        'last_date': last_date,
        'first_date': df['start_time'].iloc[0],
    }


def process_site_csv(file_path):
    """Process one site CSV. Returns (site_info, list_of_battery_results)."""
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        log.error(f"Cannot read {file_path}: {e}")
        return None, []

    if 'start_time' not in df.columns or 'stop_cause' not in df.columns:
        log.warning(f"Missing required columns in {os.path.basename(file_path)}")
        return None, []

    df['start_time'] = pd.to_datetime(df['start_time'], errors='coerce')
    df = df.dropna(subset=['start_time'])
    df = df[df['stop_cause'].isin(['MAX_TIME', 'BATTERY_CAPACITY', 'BATTERY_VOLTAGE'])].copy()
    df = df.sort_values('start_time').reset_index(drop=True)

    if len(df) < 3:
        log.warning(f"Not enough data in {os.path.basename(file_path)}: {len(df)} rows")
        return None, []

    fname = os.path.basename(file_path).replace('.csv', '')
    ip = df['IP'].iloc[0] if 'IP' in df.columns else fname
    site_name = df['site_location'].iloc[0] if 'site_location' in df.columns else fname
    batt_type = df['batt_type'].iloc[0] if 'batt_type' in df.columns else 'N/A'

    site_info = {
        'file': fname,
        'ip': str(ip),
        'site_name': str(site_name),
        'batt_type': str(batt_type),
        'total_measurements': len(df),
        'date_range_start': df['start_time'].min(),
        'date_range_end': df['start_time'].max(),
    }

    # Detect batteries
    battery_results = []
    for i in range(1, 10):
        col = f'final_cap_rate{i}'
        if col not in df.columns:
            continue
        valid = df[col].dropna()
        valid = valid[valid > 0]
        if len(valid) < 3:
            continue

        result = analyze_single_battery(df, col, i)
        if result:
            result['battery_name'] = f'String{i}'
            result['battery_idx'] = i
            battery_results.append(result)

    return site_info, battery_results


def process_all_sites(data_folder):
    """Process all CSV files in folder. Returns (all_sites, all_batteries)."""
    csv_files = sorted(glob.glob(os.path.join(data_folder, '*.csv')))
    log.info(f"Found {len(csv_files)} CSV files in {data_folder}")

    all_sites = []
    all_batteries = []

    for i, fp in enumerate(csv_files):
        log.info(f"[{i+1}/{len(csv_files)}] {os.path.basename(fp)}")
        site_info, batt_results = process_site_csv(fp)
        if site_info is None:
            continue
        all_sites.append(site_info)
        for br in batt_results:
            br['ip'] = site_info['ip']
            br['site_name'] = site_info['site_name']
            br['file'] = site_info['file']
            br['batt_type'] = site_info['batt_type']
            all_batteries.append(br)

    log.info(f"Processed {len(all_sites)} sites, {len(all_batteries)} battery strings")
    return all_sites, all_batteries


# ═══════════════════════════════════════════════════════════
# Excel Report Generation
# ═══════════════════════════════════════════════════════════

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_data_cell(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER):
    cell.font = font
    cell.alignment = alignment
    cell.border = THIN_BORDER


def get_status_fill(status):
    return {'OK': FILL_OK, 'WATCH': FILL_WATCH, 'WARNING': FILL_WARNING, 'CRITICAL': FILL_CRITICAL}.get(status, FILL_WHITE)


def auto_width(ws, min_width=10, max_width=30):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


def create_summary_sheet(wb, all_batteries):
    """Main dashboard sheet with all batteries."""
    ws = wb.active
    ws.title = 'Dashboard'
    ws.sheet_properties.tabColor = '1F4E79'

    # Title
    ws.merge_cells('A1:P1')
    ws['A1'] = 'Battery Fleet Health Dashboard'
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:P2')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}  |  Sites: {len(set(b["ip"] for b in all_batteries))}  |  Batteries: {len(all_batteries)}'
    ws['A2'].font = Font(name='Arial', size=10, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    # Status summary counts
    status_counts = {'OK': 0, 'WATCH': 0, 'WARNING': 0, 'CRITICAL': 0}
    for b in all_batteries:
        status_counts[b['status']] = status_counts.get(b['status'], 0) + 1

    row = 3
    col = 1
    for status, count in status_counts.items():
        ws.cell(row=row, column=col, value=f'{status}: {count}')
        ws.cell(row=row, column=col).font = Font(name='Arial', bold=True, size=11)
        ws.cell(row=row, column=col).fill = get_status_fill(status)
        ws.cell(row=row, column=col).alignment = ALIGN_CENTER
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        col += 2
    ws.row_dimensions[3].height = 28

    # Headers
    headers = [
        '№', 'IP', 'Site Name', 'Battery', 'Type',
        'Current SOH (%)', 'Status',
        'Rate (%/cycle)', 'Rate (%/day)',
        'Replacements', 'Segment Cycles',
        'Accelerating?',
        '80% Date', 'Days to 80%',
        '60% Date', 'Days to 60%',
        'Avg V-Drop (V)', 'Last Measured',
    ]
    hrow = 5
    for c, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=c, value=h)
    style_header_row(ws, hrow, len(headers))
    ws.row_dimensions[hrow].height = 32

    # Freeze panes
    ws.freeze_panes = f'A{hrow+1}'

    # Auto-filter
    ws.auto_filter.ref = f'A{hrow}:{get_column_letter(len(headers))}{hrow + len(all_batteries)}'

    # Sort by status priority then SOH
    status_order = {'CRITICAL': 0, 'WARNING': 1, 'WATCH': 2, 'OK': 3}
    sorted_batteries = sorted(all_batteries, key=lambda b: (status_order.get(b['status'], 9), b['current_soh']))

    # Data rows
    for i, b in enumerate(sorted_batteries):
        r = hrow + 1 + i
        row_fill = FILL_LIGHT_GRAY if i % 2 == 0 else FILL_WHITE

        values = [
            i + 1,
            b['ip'],
            b['site_name'],
            b['battery_name'],
            b['batt_type'],
            round(b['current_soh'], 2),
            b['status'],
            round(b['rate_per_cycle'], 4),
            round(b['rate_per_day'], 4),
            b['n_replacements'],
            b['current_segment_cycles'],
            'Yes ⚠' if b['accelerating'] else 'No',
            b['date_to_80'].strftime('%Y-%m-%d') if b['date_to_80'] else ('Already below' if b['current_soh'] < SOH_WARNING else '—'),
            round(b['days_to_80'], 0) if b['days_to_80'] else '—',
            b['date_to_60'].strftime('%Y-%m-%d') if b['date_to_60'] else ('Already below' if b['current_soh'] < SOH_CRITICAL else '—'),
            round(b['days_to_60'], 0) if b['days_to_60'] else '—',
            round(b['avg_volt_drop'], 2) if b['avg_volt_drop'] is not None else '—',
            b['last_date'].strftime('%Y-%m-%d') if hasattr(b['last_date'], 'strftime') else str(b['last_date']),
        ]

        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell)
            if c not in (6, 7):  # Don't override status/soh colors with row stripe
                cell.fill = row_fill

        # Color SOH cell
        soh_cell = ws.cell(row=r, column=6)
        soh_cell.fill = get_status_fill(b['status'])
        soh_cell.font = Font(name='Arial', bold=True, size=10)

        # Color status cell
        status_cell = ws.cell(row=r, column=7)
        status_cell.fill = get_status_fill(b['status'])
        status_cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF' if b['status'] == 'CRITICAL' else '000000')

        # Red text for negative rates
        if b['rate_per_cycle'] < -0.5:
            ws.cell(row=r, column=8).font = Font(name='Arial', size=10, color='CC0000')
            ws.cell(row=r, column=9).font = Font(name='Arial', size=10, color='CC0000')

    auto_width(ws)
    # Fix specific columns
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['G'].width = 12


def create_replacements_sheet(wb, all_batteries):
    """Sheet showing battery replacement history."""
    ws = wb.create_sheet('Replacements')
    ws.sheet_properties.tabColor = 'E74C3C'

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Battery Replacement History'
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    headers = ['IP', 'Site Name', 'Battery', 'Replacement Date', 'Replacement #', 'Total Replacements']
    hrow = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=c, value=h)
    style_header_row(ws, hrow, len(headers))

    r = hrow + 1
    for b in all_batteries:
        if b['n_replacements'] == 0:
            continue
        for idx, rd in enumerate(b['replacement_dates'], 1):
            values = [
                b['ip'], b['site_name'], b['battery_name'],
                rd.strftime('%Y-%m-%d') if hasattr(rd, 'strftime') else str(rd),
                idx, b['n_replacements'],
            ]
            for c, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=c, value=val)
                style_data_cell(cell)
                cell.fill = FILL_LIGHT_GRAY if (r - hrow) % 2 == 0 else FILL_WHITE
            r += 1

    if r == hrow + 1:
        ws.cell(row=r, column=1, value='No replacements detected')
        ws.cell(row=r, column=1).font = Font(name='Arial', italic=True, color='999999')

    auto_width(ws)


def create_forecast_sheet(wb, all_batteries):
    """Sheet showing SOH forecasts."""
    ws = wb.create_sheet('Forecasts')
    ws.sheet_properties.tabColor = '2ECC71'

    ws.merge_cells('A1:N1')
    ws['A1'] = f'SOH Forecasts (Next {FORECAST_HORIZON} cycles, {MEASUREMENT_INTERVAL_DAYS}-day intervals)'
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    # Dynamic headers
    headers = ['IP', 'Site Name', 'Battery', 'Current SOH']
    for i in range(1, FORECAST_HORIZON + 1):
        headers.append(f'Cycle +{i}')

    hrow = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=c, value=h)
    style_header_row(ws, hrow, len(headers))

    r = hrow + 1
    for b in all_batteries:
        if not b['forecast_soh']:
            continue
        ws.cell(row=r, column=1, value=b['ip'])
        ws.cell(row=r, column=2, value=b['site_name'])
        ws.cell(row=r, column=3, value=b['battery_name'])
        ws.cell(row=r, column=4, value=round(b['current_soh'], 2))

        for fi, fsoh in enumerate(b['forecast_soh']):
            cell = ws.cell(row=r, column=5 + fi, value=round(fsoh, 2))
            style_data_cell(cell)
            # Color code forecast values
            if fsoh < SOH_CRITICAL:
                cell.fill = FILL_CRITICAL
            elif fsoh < SOH_WARNING:
                cell.fill = FILL_WARNING
            elif fsoh < SOH_WARNING + 5:
                cell.fill = FILL_WATCH
            else:
                cell.fill = FILL_OK

        for c in range(1, 5):
            style_data_cell(ws.cell(row=r, column=c))

        r += 1

    auto_width(ws)


def create_critical_alert_sheet(wb, all_batteries):
    """Sheet with only WARNING and CRITICAL batteries."""
    ws = wb.create_sheet('⚠ Alerts')
    ws.sheet_properties.tabColor = 'FF0000'

    ws.merge_cells('A1:J1')
    ws['A1'] = '⚠ Batteries Requiring Attention'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='CC0000')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30

    alerts = [b for b in all_batteries if b['status'] in ('WARNING', 'CRITICAL')]
    alerts.sort(key=lambda b: b['current_soh'])

    headers = ['IP', 'Site Name', 'Battery', 'SOH (%)', 'Status',
               'Rate (%/cyc)', 'Est. 80% Date', 'Est. 60% Date', 'Replacements', 'Action']
    hrow = 3
    for c, h in enumerate(headers, 1):
        ws.cell(row=hrow, column=c, value=h)
    style_header_row(ws, hrow, len(headers))

    r = hrow + 1
    for b in alerts:
        if b['current_soh'] < SOH_CRITICAL:
            action = 'REPLACE IMMEDIATELY'
        elif b['current_soh'] < 70:
            action = 'Schedule replacement soon'
        else:
            action = 'Monitor closely'

        values = [
            b['ip'], b['site_name'], b['battery_name'],
            round(b['current_soh'], 2), b['status'],
            round(b['rate_per_cycle'], 4),
            b['date_to_80'].strftime('%Y-%m-%d') if b['date_to_80'] else 'Already below',
            b['date_to_60'].strftime('%Y-%m-%d') if b['date_to_60'] else ('Already below' if b['current_soh'] < SOH_CRITICAL else '—'),
            b['n_replacements'],
            action,
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_data_cell(cell)

        ws.cell(row=r, column=5).fill = get_status_fill(b['status'])
        ws.cell(row=r, column=5).font = Font(name='Arial', bold=True, size=10,
                                              color='FFFFFF' if b['status'] == 'CRITICAL' else '000000')
        ws.cell(row=r, column=4).fill = get_status_fill(b['status'])

        if b['status'] == 'CRITICAL':
            ws.cell(row=r, column=10).font = Font(name='Arial', bold=True, size=10, color='CC0000')

        r += 1

    if not alerts:
        ws.cell(row=r, column=1, value='✓ No batteries in WARNING or CRITICAL state')
        ws.cell(row=r, column=1).font = Font(name='Arial', italic=True, size=11, color='27AE60')

    auto_width(ws)
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['J'].width = 25


def generate_excel_report(all_batteries, output_path):
    """Generate the complete Excel report."""
    wb = Workbook()

    log.info("Creating Dashboard sheet...")
    create_summary_sheet(wb, all_batteries)

    log.info("Creating Alerts sheet...")
    create_critical_alert_sheet(wb, all_batteries)

    log.info("Creating Forecasts sheet...")
    create_forecast_sheet(wb, all_batteries)

    log.info("Creating Replacements sheet...")
    create_replacements_sheet(wb, all_batteries)

    wb.save(output_path)
    log.info(f"Excel report saved: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════

if __name__ == '__main__':
    DATA_FOLDER = '../data'
    OUTPUT_FILE = 'battery_fleet_report.xlsx'

    all_sites, all_batteries = process_all_sites(DATA_FOLDER)

    if all_batteries:
        generate_excel_report(all_batteries, OUTPUT_FILE)

        # Print summary
        print(f"\n{'='*70}")
        print(f"BATTERY FLEET SUMMARY")
        print(f"{'='*70}")
        for b in sorted(all_batteries, key=lambda x: x['current_soh']):
            print(f"  {b['ip']:>20s} | {b['battery_name']:>8s} | "
                  f"SOH: {b['current_soh']:6.2f}% | "
                  f"Rate: {b['rate_per_cycle']:+.4f}%/cyc | "
                  f"Status: {b['status']}")
    else:
        print("No battery data found!")