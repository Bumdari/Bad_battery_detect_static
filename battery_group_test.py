import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# ТОХИРГОО
# ─────────────────────────────────────────
DATA_FOLDER = 'data'
OUTPUT_FILE = 'battery_group_analysis.xlsx'
CAP_THRESHOLD = 65     
MIN_DURATION = 15        
BASELINE_TOP_N = 10     
SLOPE_DAYS = 30          


def get_group_cap(row, group):
    """Group-ийн final cap rate дундажийг тооцно (0.0 утгыг хасна)"""
    if group == 1:
        vals = [row['final_cap_rate1'], row['final_cap_rate2']]
    else:
        vals = [row['final_cap_rate3'], row['final_cap_rate4']]
    valid = [v for v in vals if v > 0]
    return round(np.mean(valid), 2) if valid else np.nan

def group_exists(row, group):
    """Group байгаа эсэхийг шалгана (init_cap_rate > 0)"""
    if group == 1:
        return row['init_cap_rate1'] > 0 or row['init_cap_rate2'] > 0
    else:
        return row['init_cap_rate3'] > 0 or row['init_cap_rate4'] > 0

def determine_weak_group(row):
    """
    Weak group тодорхойлох логик:
    - BATTERY_VOLTAGE / MAX_TIME: final_cap бага нь weak
    - BATTERY_CAPACITY: CAP_THRESHOLD-д хүрсэн нь weak
    """
    g1_exists = group_exists(row, 1)
    g2_exists = group_exists(row, 2)
    g1_cap = row['group1_cap']
    g2_cap = row['group2_cap']
    cause = row['stop_cause']

    if g1_exists and not g2_exists:
        return 'single_group1'
    if g2_exists and not g1_exists:
        return 'single_group2'

    if cause in ['BATTERY_VOLTAGE', 'MAX_TIME']:
        if pd.isna(g1_cap) or pd.isna(g2_cap):
            return 'unknown'
        if g1_cap < g2_cap:
            return 'group1'
        elif g2_cap < g1_cap:
            return 'group2'
        else:
            return 'equal'

    elif cause == 'BATTERY_CAPACITY':
        g1_weak = (not pd.isna(g1_cap)) and (g1_cap <= CAP_THRESHOLD)
        g2_weak = (not pd.isna(g2_cap)) and (g2_cap <= CAP_THRESHOLD)
        if g1_weak and not g2_weak:
            return 'group1'
        elif g2_weak and not g1_weak:
            return 'group2'
        elif g1_weak and g2_weak:
            return 'equal'
        else:
            if g1_cap < g2_cap:
                return 'group1'
            elif g2_cap < g1_cap:
                return 'group2'
            else:
                return 'equal'
    else:
        return 'unknown'

def is_valid_cycle(row):
    """
    Хүчинтэй cycle эсэхийг шалгана:
    - duration < MIN_DURATION ба stop_cause != BATTERY_VOLTAGE → хасах
    """
    if row['duration_min'] < MIN_DURATION and row['stop_cause'] != 'BATTERY_VOLTAGE':
        return False
    return True

def detect_battery_changes(durations):
    """Duration огцом нэмэгдэхийг battery change гэж илрүүлнэ"""
    indices = [0]
    for i in range(1, len(durations)):
        if i >= 5 and i + 5 <= len(durations):
            pre_med = durations.iloc[i-5:i].median()
            post_med = durations.iloc[i:i+5].median()
            if pre_med > 0 and post_med / pre_med >= 1.5 and post_med >= 60:
                indices.append(i)
    return indices

def calc_group_stats(df, cap_col, battery_change_idx):
    """
    Group-ийн baseline, slope, drop% тооцно
    - baseline: сүүлийн баттерийн TOP N cap median
    - slope: сүүлийн 30 хоногийн cap median
    """
    recent_df = df.iloc[battery_change_idx:].copy()
    valid_recent = recent_df[recent_df[cap_col] > 0][cap_col]

    if len(valid_recent) == 0:
        return np.nan, np.nan, np.nan

    top_n = valid_recent.sort_values(ascending=False).head(BASELINE_TOP_N)
    baseline = round(top_n.median(), 2)

    cutoff = df['start_time'].max() - pd.Timedelta(days=SLOPE_DAYS)
    last_30 = df[df['start_time'] >= cutoff]
    valid_30 = last_30[last_30[cap_col] > 0][cap_col]
    slope = round(valid_30.median(), 2) if len(valid_30) > 0 else baseline

    drop_pct = round((baseline - slope) / baseline * 100, 2) if baseline > 0 else 0.0
    return baseline, slope, drop_pct


all_files = [f for f in os.listdir(DATA_FOLDER) if f.lower().endswith('.csv')]
print(f"{len(all_files)} ширхэг CSV файл олдлоо\n")

results = []

for filename in all_files:
    filepath = os.path.join(DATA_FOLDER, filename)
    try:
        df = pd.read_csv(filepath)
        df['start_time'] = pd.to_datetime(df['start_time'])
        df = df.sort_values('start_time').reset_index(drop=True)

        valid_causes = ['BATTERY_VOLTAGE', 'BATTERY_CAPACITY', 'MAX_TIME']
        df = df[df['stop_cause'].isin(valid_causes)].reset_index(drop=True)

        if len(df) < 1:
            print(f" → {filename}: хоосон, алгасав")
            continue

        df['group1_cap'] = df.apply(lambda r: get_group_cap(r, 1), axis=1)
        df['group2_cap'] = df.apply(lambda r: get_group_cap(r, 2), axis=1)

        before = len(df)
        df = df[df.apply(is_valid_cycle, axis=1)].reset_index(drop=True)
        removed = before - len(df)
        if removed > 0:
            print(f"   ⚠ {filename}: {removed} хүчингүй cycle хасагдлаа")

        if len(df) < 1:
            print(f" → {filename}: шүүлтийн дараа хоосон, алгасав")
            continue

        change_indices = detect_battery_changes(df['duration_min'])
        last_change_idx = change_indices[-1]
        battery_change_date = df.loc[last_change_idx, 'start_time']

        df['weak_group'] = df.apply(determine_weak_group, axis=1)

        last = df.iloc[-1]
        ip = df['IP'].iloc[0]
        last_start_time = last['start_time'].strftime('%Y-%m-%d %H:%M:%S')
        last_duration = last['duration_min']
        last_stop_cause = last['stop_cause']
        last_weak = last['weak_group']

        g1_baseline, g1_slope, g1_drop = calc_group_stats(df, 'group1_cap', last_change_idx)

        has_g2 = df.apply(lambda r: group_exists(r, 2), axis=1).any()
        if has_g2:
            g2_baseline, g2_slope, g2_drop = calc_group_stats(df, 'group2_cap', last_change_idx)
        else:
            g2_baseline, g2_slope, g2_drop = np.nan, np.nan, np.nan

        cutoff = df['start_time'].max() - pd.Timedelta(days=SLOPE_DAYS)
        last_30_weak = df[df['start_time'] >= cutoff]['weak_group'].value_counts().to_dict()
        dominant_weak = max(last_30_weak, key=last_30_weak.get) if last_30_weak else 'unknown'

        results.append({
            'ip_address': ip,
            'battery_change_date': battery_change_date.strftime('%Y-%m-%d'),
            'last_start_time': last_start_time,
            'last_duration_min': last_duration,
            'last_stop_cause': last_stop_cause,
            'last_weak_group': last_weak,
            'dominant_weak_30d': dominant_weak,
            'group1_baseline': g1_baseline,
            'group1_slope': g1_slope,
            'group1_drop%': g1_drop,
            'group2_baseline': g2_baseline if has_g2 else 'N/A',
            'group2_slope': g2_slope if has_g2 else 'N/A',
            'group2_drop%': g2_drop if has_g2 else 'N/A',
        })

        print(f" ✓ {filename} → {ip} | weak: {last_weak} | g1_drop: {g1_drop}% | g2_drop: {g2_drop if has_g2 else 'N/A'}%")

    except Exception as e:
        print(f" ✗ {filename} алдаа: {e}")


if not results:
    print("\nЯмар ч файл боловсруулж чадсангүй!")
else:
    out_df = pd.DataFrame(results)
    out_df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Баганын индекс олох
    col_map = {}
    for c in range(1, ws.max_column + 1):
        col_map[ws.cell(1, c).value] = c

    # Өнгөний тодорхойлолт
    red    = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
    orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    yellow = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    green  = PatternFill(start_color='00CC66', end_color='00CC66', fill_type='solid')
    gray   = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    for row in range(2, ws.max_row + 1):

        # weak_group өнгө
        for col_name in ['last_weak_group', 'dominant_weak_30d']:
            if col_name in col_map:
                val = ws.cell(row, col_map[col_name]).value
                if val == 'group1':
                    ws.cell(row, col_map[col_name]).fill = orange
                elif val == 'group2':
                    ws.cell(row, col_map[col_name]).fill = yellow
                elif val == 'equal':
                    ws.cell(row, col_map[col_name]).fill = red
                elif val and val.startswith('single'):
                    ws.cell(row, col_map[col_name]).fill = gray

        # drop% өнгө
        for drop_col in ['group1_drop%', 'group2_drop%']:
            if drop_col in col_map:
                val = ws.cell(row, col_map[drop_col]).value
                if isinstance(val, (int, float)):
                    if val < 25:
                        ws.cell(row, col_map[drop_col]).fill = green
                    elif val <= 65:
                        ws.cell(row, col_map[drop_col]).fill = orange
                    else:
                        ws.cell(row, col_map[drop_col]).fill = red

    # Header өнгө
    header_fill = PatternFill(start_color='2F4F8F', end_color='2F4F8F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = header_fill
        ws.cell(1, c).font = header_font
        ws.cell(1, c).alignment = Alignment(horizontal='center')

    # Багана өргөн
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22

    wb.save(OUTPUT_FILE)

    print(f"\n{'='*50}")
    print(f"ДУУСЛАА: {len(results)} IP боловсруулагдлаа")
    print(f"Файл: {OUTPUT_FILE}")
    print(f"\nОнцлох баганууд:")
    print(f"  last_weak_group   → сүүлийн cycle-д аль group сул байсан")
    print(f"  dominant_weak_30d → сүүлийн 30 хоногт давамгайлж сул байсан group")
    print(f"  group1/2_drop%    → cap буурсан хувь (NORMAL<25 / STRONG<65 / CRITICAL)")