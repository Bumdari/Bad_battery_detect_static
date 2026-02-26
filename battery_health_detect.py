"""
BATTERY HEALTH DETECTION PIPELINE v3.2 (Засварласан)
══════════════════════════════════════════════════
- Changepoint-д low dr noise бууруулах нэмсэн (min_dr_threshold)
- Duration <30 check-д cap_drop >5% шаардлага нэмсэн (идэвхтэй цэнэглэгдэх батерейд л)
- Base None бол "UNKNOWN" статус нэмсэн
- CLR, RANK, descs-д UNKNOWN нэмсэн
"""

import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CONFIG
CFG = {
    "cp_drop":      0.40,
    "cp_stable_n":     5,
    "cp_ratio":     0.80,
    "window_days":    30,
    "warn_x":       1.30,
    "crit_x":       1.60,
    "slope_weeks":     3,
    "min_dr_threshold": 0.05,  # New: low dr noise ignore
    "min_cap_drop":     5,     # New: for duration check
    "skip": {"all_data_from_DB.csv"},
    "need_cols": {"start_cause","stop_cause","start_time","duration_min","init_batt_volt1"},
}
RANK = {"CRITICAL":3,"DEGRADING":2,"WARNING":1,"NORMAL":0, "UNKNOWN":-1}

# АЛХАМ 1: ЦЭВЭРЛЭХ
def step1(df):
    df = df.copy()
    df["start_time"] = pd.to_datetime(df["start_time"])
    df = df.sort_values("start_time").reset_index(drop=True)
    df = df[df["start_cause"] == "SCHEDULED"].reset_index(drop=True)
    df = df[df["stop_cause"].isin(["MAX_TIME","BATTERY_VOLTAGE","BATTERY_CAPACITY"])].reset_index(drop=True)
    for s in [1, 2]:
        active = df[f"init_batt_volt{s}"] > 0
        df[f"active{s}"] = active
        drop = df[f"init_cap_rate{s}"] - df[f"final_cap_rate{s}"]
        dur  = df["duration_min"].replace(0, np.nan)
        df[f"dr{s}"] = np.where(active & dur.notna(), drop / dur, np.nan)
    return df

# АЛХАМ 2: CHANGEPOINT (Low dr noise бууруулах нэмсэн)
def step2(df, s):
    """Changepoint -> (df_with_segments, {seg_id: "огноо"})"""
    df   = df.copy()
    dr   = f"dr{s}"
    seg  = f"segment{s}"
    idx  = df.index[df[dr].notna()].tolist()
    sarr = np.zeros(len(df), dtype=int)
    cp_dates = {}
    if len(idx) < CFG["cp_stable_n"] + 2:
        df[seg] = 0
        return df, cp_dates
    sid, base, cand, stable, cand_i = 0, None, None, 0, None
    for i in idx:
        v = df.at[i, dr]
        if base is None:
            base = v; sarr[i] = sid; continue
        if base < CFG["min_dr_threshold"] or v < CFG["min_dr_threshold"]:
            drop_r = 0  # Low dr - no change considered
        else:
            drop_r = (base - v) / base if base > 0 else 0
        if cand is None and abs(drop_r) >= CFG["cp_drop"]:  # Бууралт эсвэл өсөлт
            cand = v; stable = 1; cand_i = i
        elif cand is not None:
            if abs(v - cand) <= abs(cand) * (1 - CFG["cp_ratio"]):
                stable += 1
            else:
                cand = None; stable = 0; cand_i = None
            if stable >= CFG["cp_stable_n"]:
                sid += 1
                cp_dates[sid] = str(df.at[cand_i, "start_time"].date())
                base = cand; cand = None; stable = 0; cand_i = None
        sarr[i] = sid
    df[seg] = sarr
    return df, cp_dates

# АЛХАМ 3: SEGMENT MEDIAN (Fallback нэмсэн)
def step3(df, s):
    dr   = f"dr{s}"
    seg  = f"segment{s}"
    meds = {}
    for sid, grp in df.groupby(seg):
        mt = grp[grp["stop_cause"] == "MAX_TIME"][dr].dropna()
        if len(mt) >= 3:
            meds[sid] = mt.median()
        else:
            long = grp[grp["duration_min"] >= 60][dr].dropna()
            if len(long) >= 2:
                meds[sid] = long.median()
            elif len(grp[dr].dropna()) >= 2:
                meds[sid] = grp[dr].dropna().median()
            else:
                meds[sid] = None  # Мэдээлэл хүрэлцэхгүй
    return meds

# АЛХАМ 4: МУУДАЛТ ( <30 мин дундаж + cap_drop >5% бол CRITICAL )
def step4(df, s, meds):
    df    = df.copy()
    dr    = f"dr{s}"
    seg   = f"segment{s}"
    stcol = f"status{s}"
    rcol  = f"roll{s}"
    df[stcol] = "NORMAL"

    # Сүүлчийн сегментийн дундаж duration, cap_drop шалгах
    lseg = int(df[seg].max())
    last_seg = df[df[seg] == lseg]
    avg_dur = last_seg["duration_min"].mean()
    avg_drop = last_seg[f"init_cap_rate{s}"].mean() - last_seg[f"final_cap_rate{s}"].mean()
    if avg_dur < 30 and avg_drop > CFG["min_cap_drop"]:
        df.loc[df[seg] == lseg, stcol] = "CRITICAL"
        return df, False  # Slope шаардлагагүй

    df = df.sort_values("start_time").reset_index(drop=True)
    roll = []
    for i, row in df.iterrows():
        cut = row["start_time"] - pd.Timedelta(days=CFG["window_days"])
        w   = df.loc[(df["start_time"] >= cut) & (df.index <= i), dr].dropna()
        roll.append(w.median() if len(w) >= 3 else np.nan)
    df[rcol] = roll

    # Slope
    wk = df.set_index("start_time")[dr].resample("7D").median().dropna()
    slope = (len(wk) >= CFG["slope_weeks"] and
             all(wk.iloc[-CFG["slope_weeks"]:].diff().dropna() > 0))

    for i, row in df.iterrows():
        base = meds.get(row[seg])
        if base is None:
            df.at[i, stcol] = "UNKNOWN"
            continue
        rm   = row[rcol]
        if pd.isna(rm) or base == 0:
            continue
        r = rm / base
        if r >= CFG["crit_x"]:
            df.at[i, stcol] = "CRITICAL"
        elif r >= CFG["warn_x"] and slope:
            df.at[i, stcol] = "DEGRADING"
        elif r >= CFG["warn_x"]:
            df.at[i, stcol] = "WARNING"
    return df, slope

# НЭГ SITE БОЛОВСРУУЛАХ (dr дундаж нэмсэн)
def process(df_raw):
    site    = df_raw["site_location"].iloc[0] if "site_location" in df_raw.columns else "Unknown"
    site_ip = df_raw["IP"].iloc[0] if "IP" in df_raw.columns else "—"
    res     = {"site": site, "site_ip": site_ip, "strings": {}, "overall": "NORMAL"}

    df = step1(df_raw)
    if len(df) < 5:
        res["error"] = f"Session хэт цөөн ({len(df)})"
        return res

    for s in [1, 2]:
        if not df[f"active{s}"].any():
            continue
        df, cp_dates  = step2(df, s)
        meds          = step3(df, s)
        df, slope     = step4(df, s, meds)

        lseg = int(df[f"segment{s}"].max())
        l30  = df[
            (df[f"segment{s}"] == lseg) &
            (df["start_time"] >= df["start_time"].max() - pd.Timedelta(days=30))
        ]
        st = "NORMAL"
        for lvl in ["CRITICAL","DEGRADING","WARNING","UNKNOWN"]:
            if (l30[f"status{s}"] == lvl).any():
                st = lvl; break

        base    = meds.get(lseg)
        last_dr = df[f"dr{s}"].dropna().iloc[-1] if df[f"dr{s}"].notna().any() else None
        ratio   = round(last_dr / base, 2) if last_dr and base else None

        cp_str = ", ".join(cp_dates[k] for k in sorted(cp_dates)) if cp_dates else "—"

        res["strings"][s] = {
            "status":   st,
            "segments": lseg + 1,
            "baseline": round(base, 3)    if base    else None,
            "last_dr":  round(last_dr, 3) if last_dr else None,
            "ratio":    ratio,
            "slope":    slope,
            "sessions": len(df),
            "date":     str(df["start_time"].max().date()),
            "btype":    df_raw["batt_type"].iloc[0] if "batt_type" in df_raw.columns else "—",
            "cp_dates": cp_str,
        }

    # Хоёр батерейны last_dr дундажлах
    if 1 in res["strings"] and 2 in res["strings"]:
        ld1 = res["strings"][1]["last_dr"]
        ld2 = res["strings"][2]["last_dr"]
        if ld1 and ld2:
            res["avg_last_dr"] = round((ld1 + ld2) / 2, 3)

    res["overall"] = max(
        (v["status"] for v in res["strings"].values()),
        key=lambda x: RANK.get(x, 0), default="NORMAL"
    )
    return res

# БҮХ SITE УНШИХ (таны folder-оос)
def run_all(folder):
    files = sorted([f for f in Path(folder).glob("*.csv") if f.name not in CFG["skip"]])
    print(f"📂 {len(files)} CSV файл\n{'─'*62}")
    out, skip = [], 0
    for i, fp in enumerate(files, 1):
        try:
            df = pd.read_csv(str(fp))
            if CFG["need_cols"] - set(df.columns):
                print(f"  {i:>3}. ⏭️  {fp.name:<35} форматгүй — алгасав")
                skip += 1; continue
            r    = process(df)
            icon = {"CRITICAL":"🔴","DEGRADING":"🟡","WARNING":"🟠","NORMAL":"✅","UNKNOWN":"⚪"}.get(r["overall"],"❓")
            print(f"  {i:>3}. {icon} {r['site'][:54]:<54} {r['overall']}")
            out.append(r)
        except Exception as e:
            print(f"  {i:>3}. ⚠️  {fp.name} — {e}")
    print(f"\n  ✔ Боловсруулсан: {len(out)}   Алгасав: {skip}")
    return out

# EXCEL EXPORT (таны формат хэвээр, UNKNOWN нэмсэн)
CLR = {
    "CRITICAL":   ("FF4444", "FFFFFF"),   # Улаан
    "DEGRADING":  ("FF8C00", "FFFFFF"),   # Улбар шар
    "WARNING":    ("FFD700", "000000"),   # Шар
    "NORMAL":     ("4CAF50", "FFFFFF"),   # Ногоон
    "UNKNOWN":    ("A9A9A9", "000000"),   # Саарал - New
    "HDR":        ("1F3864", "FFFFFF"),   # Харанхуй цэнхэр — гарчиг
    "ODD":        ("FFFFFF", "000000"),   # Цагаан мөр
    "EVEN":       ("EBF3FB", "000000"),   # Цайвар цэнхэр мөр
    "RATIO_CRIT": ("FFCCCC", "000000"),   # Ratio >= 1.6
    "RATIO_WARN": ("FFE599", "000000"),   # Ratio >= 1.3
    "SLOPE_UP":   ("FF4444", "FFFFFF"),   # Slope өсч байна — улаан
    "NONE":       ("F2F2F2", "A0A0A0"),   # Мэдээлэл байхгүй
}

COL_WIDTHS = [4, 15, 46, 14, 13, 12, 11, 10, 9, 11, 20, 13, 12, 11, 10, 9, 11, 20, 13, 10, 15]

COL_HEADERS = [
    "№", "Site IP", "Site нэр", "Нийт байдал",
    "Battery 1", "Battery 1\nBaseline DR", "Battery 1\nLast DR",
    "Battery 1\nRatio", "Battery 1\nSlope", "Battery 1\nСолигдсон тоо",
    "Battery 1\nШинэ battery\nогноо",
    "Battery 2", "Battery 2\nBaseline DR", "Battery 2\nLast DR",
    "Battery 2\nRatio", "Battery 2\nSlope", "Battery 2\nСолигдсон тоо",
    "Battery 2\nШинэ battery\nогноо",
    "Batt Type", "Sessions", "Сүүлийн огноо",
]

def fl(bg):             return PatternFill("solid", fgColor=bg)
def ft(fg, bold=False, sz=10):
    return Font(color=fg, bold=bold, size=sz, name="Calibri")
def bd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def al(h="center"):     return Alignment(horizontal=h, vertical="center", wrap_text=True)

def make_summary(ws, results):
    for ci, (hdr, w) in enumerate(zip(COL_HEADERS, COL_WIDTHS), 1):
        c = ws.cell(1, ci, hdr)
        c.fill      = fl(CLR["HDR"][0])
        c.font      = ft(CLR["HDR"][1], bold=True, sz=10)
        c.alignment = al()
        c.border    = bd()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32

    srt = sorted(results, key=lambda x: -RANK.get(x["overall"], 0))

    for ri, r in enumerate(srt, 2):
        s1  = r["strings"].get(1, {})
        s2  = r["strings"].get(2, {})
        bg  = CLR["EVEN"][0] if ri % 2 == 0 else CLR["ODD"][0]

        s1_slope = "Муудаж эхэлж байна" if s1.get("slope") else "—"
        s2_slope = "Муудаж эхэлж байна" if s2.get("slope") else "—"

        vals = [
            ri - 1,
            r.get("site_ip", "—"),
            r["site"],
            r["overall"],
            s1.get("status",   "—"),
            s1.get("baseline", "—"),
            s1.get("last_dr",  "—"),
            s1.get("ratio",    "—"),
            s1_slope,
            s1.get("segments", "—"),
            s1.get("cp_dates", "—"),
            s2.get("status",   "—"),
            s2.get("baseline", "—"),
            s2.get("last_dr",  "—"),
            s2.get("ratio",    "—"),
            s2_slope,
            s2.get("segments", "—"),
            s2.get("cp_dates", "—"),
            s1.get("btype",    s2.get("btype",    "—")),
            s1.get("sessions", s2.get("sessions", "—")),
            s1.get("date",     s2.get("date",     "—")),
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(ri, ci, v if v is not None else "—")
            c.border    = bd()
            c.alignment = al() if ci != 2 else al("left")

            if v is None:
                c.fill = fl(CLR["NONE"][0])
                c.font = ft(CLR["NONE"][1], sz=10)

            elif ci == 4 and v in CLR:
                c.fill = fl(CLR[v][0])
                c.font = ft(CLR[v][1], bold=True, sz=10)

            elif ci in (5, 12) and v in CLR:
                c.fill = fl(CLR[v][0])
                c.font = ft(CLR[v][1], sz=10)

            elif ci in (8, 15):
                try:
                    fv = float(v)
                    if fv >= CFG["crit_x"]:
                        c.fill = fl(CLR["RATIO_CRIT"][0])
                        c.font = ft(CLR["RATIO_CRIT"][1], bold=True, sz=10)
                    elif fv >= CFG["warn_x"]:
                        c.fill = fl(CLR["RATIO_WARN"][0])
                        c.font = ft(CLR["RATIO_WARN"][1], sz=10)
                    else:
                        c.fill = fl(bg)
                        c.font = ft("000000", sz=10)
                except:
                    c.fill = fl(CLR["NONE"][0])
                    c.font = ft(CLR["NONE"][1], sz=10)

            elif ci in (9, 16):
                if v == "Муудаж эхэлж байна":
                    c.fill = fl(CLR["SLOPE_UP"][0])
                    c.font = ft(CLR["SLOPE_UP"][1], bold=True, sz=10)
                else:
                    c.fill = fl(bg)
                    c.font = ft("A0A0A0", sz=10)

            elif ci in (11, 18):
                if v != "—":
                    c.fill = fl("FFF2CC")
                    c.font = ft("7F6000", sz=9)
                else:
                    c.fill = fl(bg)
                    c.font = ft("A0A0A0", sz=10)

            else:
                c.fill = fl(bg)
                c.font = ft("000000", sz=10)

        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(21)}1"

def make_stats(ws, results):
    total  = len(results)
    counts = {k: 0 for k in ["CRITICAL","DEGRADING","WARNING","NORMAL","UNKNOWN"]}
    for r in results:
        counts[r["overall"]] = counts.get(r["overall"], 0) + 1

    ws.merge_cells("A1:E1")
    t = ws.cell(1, 1, "⚡  BATTERY HEALTH — ДҮГНЭЛТ")
    t.fill = fl(CLR["HDR"][0]); t.font = ft(CLR["HDR"][1], bold=True, sz=13)
    t.alignment = al(); ws.row_dimensions[1].height = 32

    for ci, h in enumerate(["Статус","Site тоо","Хувь","Үнэлгээ","Тайлбар"], 1):
        c = ws.cell(2, ci, h)
        c.fill = fl("2E75B6"); c.font = ft("FFFFFF", bold=True)
        c.alignment = al(); c.border = bd()
    ws.row_dimensions[2].height = 24

    descs = {
        "CRITICAL":  ("🔴 Яаралтай",    "Даруй солих шаардлагатай — хойшлуулах боломжгүй"),
        "DEGRADING": ("🟡 Муудаж байна", "Тасралтгүй доройтож байна — удахгүй солих"),
        "WARNING":   ("🟠 Анхааруулга",  "Хэвийн хэмжээнээс хэтэрч байна — хянах"),
        "NORMAL":    ("✅ Хэвийн",        "Хэвийн ажиллагаатай"),
        "UNKNOWN":   ("⚪ Мэдээлэл дутуу", "Мэдээлэл хангалтгүй — шинжилгээ хийх боломжгүй"),  # New
    }
    for ri, st in enumerate(["CRITICAL","DEGRADING","WARNING","NORMAL","UNKNOWN"], 3):
        cnt = counts[st]
        pct = round(cnt / total * 100, 1) if total else 0
        icon_lbl, desc = descs.get(st, ("—", "—"))
        for ci, v in enumerate([st, cnt, f"{pct}%", icon_lbl, desc], 1):
            c = ws.cell(ri, ci, v)
            c.fill = fl(CLR.get(st, CLR["NONE"])[0]); c.font = ft(CLR.get(st, CLR["NONE"])[1], bold=(ci==1), sz=11)
            c.alignment = al(); c.border = bd()
        ws.row_dimensions[ri].height = 24

    for ci, v in enumerate(["НИЙТ", total, "100%"], 1):
        c = ws.cell(8, ci, v)  # Adjusted row
        c.font = ft("000000", bold=True, sz=11)
        c.border = bd(); c.alignment = al()
    ws.row_dimensions[8].height = 24

    ws.cell(10, 1, "📌 Нэр томьёо").font = ft("1F3864", bold=True, sz=12)
    notes = [
        ("Baseline DR",   "MAX_TIME session-үүдийн discharge rate-ийн median — хэвийн ажиллагааны лавлах утга"),
        ("Last DR",       "Хамгийн сүүлийн session-ий discharge rate — одоогийн байдал"),
        ("Ratio",         "Last DR ÷ Baseline  |  1.0=хэвийн  |  1.3+=анхааруулга  |  1.6+=критикал"),
        ("↑ Муу (Slope)", "Сүүлийн 3 долоо хоног дараалан DR өсч байвал — муудаж байгааг илэрхийлнэ"),
        ("Солигдсон тоо", "Changepoint-ээр battery хэдэн удаа солигдсон байж болохыг тооцсон"),
        ("DR тооцоо",     "Discharge Rate = cap_drop ÷ duration_min  (минут тутам алдсан цэнэгийн %)"),
    ]
    for ri, (term, expl) in enumerate(notes, 11):
        ws.cell(ri, 1, term).font = ft("1F3864", bold=True)
        c2 = ws.cell(ri, 2, expl)
        c2.font = ft("000000"); c2.alignment = al("left")
        ws.merge_cells(f"B{ri}:E{ri}")
        ws.row_dimensions[ri].height = 20

    for ci, w in enumerate([18, 14, 10, 20, 45], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

def export_excel(results, out_path):
    print(f"\n📊 Excel үүсгэж байна...")
    wb  = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "📋 Тайлан"
    make_summary(ws1, results)
    ws2 = wb.create_sheet("📊 Статистик")
    make_stats(ws2, results)
    wb.save(out_path)
    print(f"✅ Хадгалагдлаа → {out_path}")

# MAIN
if __name__ == "__main__":
    UPLOAD = "data"  # Таны folder
    OUTPUT = "battery_health_report_3.xlsx"

    print("═" * 62)
    print("   BATTERY HEALTH DETECTION  v3.2")
    print("═" * 62)

    results = run_all(UPLOAD)

    if results:
        print(f"\n{'─'*40}")
        counts = {}
        for r in results:
            counts[r["overall"]] = counts.get(r["overall"], 0) + 1
        icons = {"CRITICAL":"🔴","DEGRADING":"🟡","WARNING":"🟠","NORMAL":"✅","UNKNOWN":"⚪"}
        for lvl in ["CRITICAL","DEGRADING","WARNING","NORMAL","UNKNOWN"]:
            if lvl in counts:
                print(f"  {icons[lvl]} {lvl:<12} {counts[lvl]:>3} site")
        print(f"  {'─'*28}\n     Нийт       {len(results):>3} site")
        export_excel(results, OUTPUT)
    else:
        print("⚠️  Боловсруулах файл олдсонгүй.")