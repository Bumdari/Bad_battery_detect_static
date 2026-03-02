"""
Battery Discharge Analysis Script
"""
import sys, os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VALID_STOP_CAUSE = ['MAX_TIME', 'BATTERY_VOLTAGE', 'BATTERY_CAPACITY']
PREV_WINDOW=10; NEXT_WINDOW=5; PREV_CAP_MAX=0.2; NEXT_CAP_MIN=0.6; MIN_GAP=5
TARGET_MEDIAN=60; STOPPER_BAD=70; STOPPER_WEAK=40; STOPPER_GOOD=15

def load_and_filter(fp):
    df = pd.read_csv(fp)
    df['start_time'] = pd.to_datetime(df['start_time'])
    df = df.sort_values('start_time').reset_index(drop=True)
    mask = (df['start_cause'] == 'SCHEDULED') & (df['stop_cause'].isin(VALID_STOP_CAUSE))
    return df[mask].reset_index(drop=True)

def detect_replacements(df):
    df = df.copy()
    df['is_cap'] = df['stop_cause'].eq('BATTERY_CAPACITY').astype(int)
    candidates = []
    for i in range(PREV_WINDOW, len(df) - NEXT_WINDOW):
        if (df['is_cap'].iloc[i-PREV_WINDOW:i].mean() < PREV_CAP_MAX and
                df['is_cap'].iloc[i:i+NEXT_WINDOW].mean() >= NEXT_CAP_MIN):
            candidates.append(i)
    cp = []
    for idx in candidates:
        if not cp or idx - cp[-1] > MIN_GAP: cp.append(idx)
    return cp

def smart_median(durations):
    med = durations.median()
    if med >= TARGET_MEDIAN: return float(med)
    return float(durations.iloc[(durations - TARGET_MEDIAN).abs().argsort().iloc[0]])

def detect_degradation(df_seg):
    seg_med = smart_median(df_seg['duration_min'])
    cutoff  = df_seg['start_time'].max() - pd.Timedelta(days=30)
    recent  = df_seg[df_seg['start_time'] >= cutoff]
    if len(recent) < 3: return 'STABLE', round(seg_med,1), None, None
    rec_med = float(recent['duration_min'].median())
    ratio   = seg_med / rec_med if rec_med > 0 else 999.0
    status  = 'CRITICAL' if ratio >= 1.6 else ('DEGRADING' if ratio >= 1.3 else 'STABLE')
    return status, round(seg_med,1), round(rec_med,1), round(ratio,2)

def analyze_strings(df_seg, active):
    counts={s:0 for s in active}; caps={s:[] for s in active}; volts={s:[] for s in active}
    total_stops=0
    for _, row in df_seg.iterrows():
        for s in active:
            caps[s].append(row[f'final_cap_rate{s}'])
            volts[s].append(row[f'final_batt_volt{s}'])
        stop = row['stop_cause']
        if stop == 'MAX_TIME': continue
        total_stops += 1
        if stop == 'BATTERY_VOLTAGE':
            stopper = min(active, key=lambda s: row[f'final_batt_volt{s}'])
        else:
            stopper = min(active, key=lambda s: row[f'final_cap_rate{s}'])
        counts[stopper] += 1
    results = {}
    for s in active:
        pct = round(counts[s]/total_stops*100, 1) if total_stops > 0 else 0.0
        results[s] = {'stopper_count': counts[s], 'stopper_pct': pct,
                      'avg_cap': round(float(np.mean(caps[s])),1),
                      'avg_volt': round(float(np.mean(volts[s])),2)}
    return results, total_stops

def classify_string(pct, avg_cap, other_results):
    cap_gap = (max(r['avg_cap'] for r in other_results) - avg_cap) if other_results else 0
    if pct >= STOPPER_BAD:
        return 'BAD',  f'{pct:.0f}% discharge → батерей солих'
    elif pct >= STOPPER_WEAK:
        return 'WEAK', f'{pct:.0f}% discharge → муудаж байна'
    elif cap_gap >= 15:
        return 'WEAK', f'Нөгөө battery-нээс cap {cap_gap:.0f}% доогуур'
    elif pct <= STOPPER_GOOD and avg_cap >= 82:
        return 'GOOD', f'Зөвхөн {pct:.0f}% тохиолдолд зогсоосон — сайн'
    else:
        return 'FAIR', f'{pct:.0f}% зогсоосон, cap={avg_cap:.0f}%'

def analyze_file(fp):
    print(f"\n {os.path.basename(fp)}")
    df = load_and_filter(fp)
    if df.empty: print("    Хоосон"); return None
    ip   = df['IP'].iloc[0] if 'IP' in df.columns else '-'
    site = df['site_location'].iloc[0] if 'site_location' in df.columns else '-'
    cp   = detect_replacements(df)
    rep_date = df.iloc[cp[-1]]['start_time'].date() if cp else None
    df_seg   = df.iloc[cp[-1]:].copy().reset_index(drop=True) if cp else df.copy()
    deg_status, seg_med, rec_med, ratio = detect_degradation(df_seg)
    active = [s for s in [1,2,3,4] if (df_seg[f'init_cap_rate{s}'] > 0).sum() > 0]
    str_results, total_stops = analyze_strings(df_seg, active)
    strings = {}
    for s in active:
        r = str_results[s]
        others = [str_results[o] for o in active if o != s]
        status, note = classify_string(r['stopper_pct'], r['avg_cap'], others)
        strings[s] = {'status': status, 'stopper_pct': r['stopper_pct'],
                      'stopper_count': r['stopper_count'],
                      'avg_cap': r['avg_cap'], 'avg_volt': r['avg_volt'], 'note': note}
    deg_icon = {'CRITICAL':'[CRITICAL]','DEGRADING':'[DEGRADING]','STABLE':'[STABLE]'}[deg_status]
    print(f"   Батерей солисон : {rep_date or 'Анхны батери'}")
    print(f"   Сегмент n={len(df_seg)}  Муудалт: {deg_icon}  ratio={ratio}  ({seg_med}мин / {rec_med}мин)")
    for s, v in strings.items():
        print(f"   String{s}: {v['status']:5s} | {v['stopper_count']:3d} удаа ({v['stopper_pct']:5.1f}%) | cap={v['avg_cap']}%  volt={v['avg_volt']}V | {v['note']}")
    return {'ip':ip,'site':site,'rep_date':str(rep_date) if rep_date else None,
            'seg_start':str(df_seg.iloc[0]['start_time'].date()),
            'seg_end':str(df_seg.iloc[-1]['start_time'].date()),
            'n_seg':len(df_seg),'total_stops':total_stops,
            'seg_med':seg_med,'rec_med':rec_med,'ratio':ratio,
            'deg_status':deg_status,'last_dur':int(df_seg.iloc[-1]['duration_min']),
            'strings':strings,'active':active}

STR_FILL = {
    'BAD':  PatternFill('solid', start_color='FF0000'),
    'WEAK': PatternFill('solid', start_color='FCE4D6'),
    'FAIR': PatternFill('solid', start_color='FFF2CC'),
    'GOOD': PatternFill('solid', start_color='E2EFDA'),
}
STR_FONT = {
    'BAD':  Font(name='Arial', size=10, bold=True, color='FFFFFF'),
    'WEAK': Font(name='Arial', size=10, bold=True, color='C00000'),
    'FAIR': Font(name='Arial', size=10, bold=True, color='7F6000'),
    'GOOD': Font(name='Arial', size=10, bold=True, color='375623'),
}
STR_LABEL = {'BAD':'✕  BAD','WEAK':'⚠  WEAK','FAIR':'~  FAIR','GOOD':'✓  GOOD'}
DEG_FILL  = {'CRITICAL':PatternFill('solid',start_color='FCE4D6'),
             'DEGRADING':PatternFill('solid',start_color='FFF2CC'),
             'STABLE':PatternFill('solid',start_color='E2EFDA')}
DEG_FONT  = {'CRITICAL':Font(name='Arial',size=10,bold=True,color='C00000'),
             'DEGRADING':Font(name='Arial',size=10,bold=True,color='7F6000'),
             'STABLE':Font(name='Arial',size=10,bold=True,color='375623')}
DEG_LABEL = {'CRITICAL':'✕  CRITICAL','DEGRADING':'⚠  DEGRADING','STABLE':'✓  STABLE'}

def sort_sites(all_sites):
    """
    Эрэмблэлт:
    1. Муудалт: CRITICAL → DEGRADING → STABLE
    2. String-ийн BAD тоо (олон BAD → дээр)
    3. Ratio буурах дарааллаар
    """
    deg_order = {'CRITICAL': 0, 'DEGRADING': 1, 'STABLE': 2}
    def sort_key(s):
        bad_count = sum(1 for v in s['strings'].values() if v['status'] == 'BAD')
        return (deg_order.get(s['deg_status'], 9), -bad_count, -(s['ratio'] or 0))
    return sorted(all_sites, key=sort_key)


def build_excel(all_sites, out_path):
    all_sites = sort_sites(all_sites)
    wb = Workbook(); ws = wb.active
    ws.title = "Battery Health Report"
    ws.sheet_view.showGridLines = False
    thin = Side(style='thin', color='CCCCCC')
    med  = Side(style='medium', color='888888')
    B    = Border(top=thin,right=thin,bottom=thin,left=thin)
    BH   = Border(top=thin,right=thin,bottom=med, left=thin)
    DARK = PatternFill('solid',start_color='1F4E79')
    MED  = PatternFill('solid',start_color='2E75B6')
    LT   = PatternFill('solid',start_color='DEEAF1')
    WH   = PatternFill('solid',start_color='FFFFFF')
    ALT  = PatternFill('solid',start_color='F5F9FD')
    SEP  = PatternFill('solid',start_color='D0D0D0')
    WH11 = Font(name='Arial',size=11,bold=True,color='FFFFFF')
    WH10 = Font(name='Arial',size=10,bold=True,color='FFFFFF')
    BL10 = Font(name='Arial',size=10,bold=True,color='1F4E79')
    F10  = Font(name='Arial',size=10)
    F10B = Font(name='Arial',size=10,bold=True)
    F9   = Font(name='Arial',size=9,color='444444')
    F9I  = Font(name='Arial',size=9,italic=True,color='999999')
    F10I = Font(name='Arial',size=10,italic=True,color='999999')
    C = Alignment(horizontal='center',vertical='center',wrap_text=True)
    L = Alignment(horizontal='left',  vertical='center',wrap_text=True)

    def sc(r,c,v,font=F10,fill=WH,aln=C,brd=B):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=font;cell.fill=fill;cell.alignment=aln;cell.border=brd
        return cell
    def mg(r1,c1,r2,c2,v='',**kw):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        sc(r1,c1,v,**kw)

    ws_widths = [16,32,16,24,7,13,14,8,16,12, 2, 16,12,10,10,32, 2, 16,12,10,10,32]
    for i,w in enumerate(ws_widths,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    mg(1,1,1,2,'SITE', font=WH11,fill=DARK,aln=C)
    mg(1,3,1,10,'БАТЕРЕЙН ЕРӨНХИЙ БАЙДАЛ', font=WH11,fill=MED,aln=C)
    sc(1,11,'',fill=SEP)
    mg(1,12,1,16,'STRING 1', font=WH11,fill=DARK,aln=C)
    sc(1,17,'',fill=SEP)
    mg(1,18,1,22,'STRING 2', font=WH11,fill=DARK,aln=C)

    ws.row_dimensions[2].height = 42
    hdr2 = [
        (1,'IP хаяг',LT,BL10),(2,'Байршил',LT,BL10),
        (3,'Батерей\nсолисон',MED,WH10),(4,'Анализын хугацаа',MED,WH10),
        (5,'n',MED,WH10),(6,'Segment\nMedian (мин)',MED,WH10),
        (7,'Сүүлийн\n30 хоног (мин)',MED,WH10),(8,'Ratio',MED,WH10),
        (9,'Муудалтын\nтүвшин',MED,WH10),(10,'Сүүлийн\ndischarge (мин)',MED,WH10),
        (11,'',SEP,F10),
        (12,'Статус',LT,BL10),(13,'Зогсоосон\n%',LT,BL10),
        (14,'Avg\nCap(%)',LT,BL10),(15,'Avg\nVolt(V)',LT,BL10),(16,'Тайлбар',LT,BL10),
        (17,'',SEP,F10),
        (18,'Статус',MED,WH10),(19,'Зогсоосон\n%',MED,WH10),
        (20,'Avg\nCap(%)',MED,WH10),(21,'Avg\nVolt(V)',MED,WH10),(22,'Тайлбар',MED,WH10),
    ]
    for col,txt,fill,font in hdr2:
        sc(2,col,txt,font=font,fill=fill,aln=C,brd=BH)

    for idx,s in enumerate(all_sites):
        row = 3+idx
        ws.row_dimensions[row].height = 22
        bg = ALT if idx%2==0 else WH
        sc(row,1,s['ip'],font=F10B,fill=bg,aln=L)
        sc(row,2,s['site'],font=F10,fill=bg,aln=L)
        sc(row,3,s['rep_date'] if s['rep_date'] else 'Анхны батери',
           font=F10 if s['rep_date'] else F10I,fill=bg,aln=C)
        sc(row,4,f"{s['seg_start']}  →  {s['seg_end']}",font=F9,fill=bg,aln=C)
        sc(row,5,s['n_seg'],font=F10,fill=bg,aln=C)
        sc(row,6,s['seg_med'],font=F10,fill=bg,aln=C)
        sc(row,7,s['rec_med'] if s['rec_med'] else 'N/A',font=F10,fill=bg,aln=C)
        sc(row,8,s['ratio'] if s['ratio'] else '-',font=F10,fill=bg,aln=C)
        ds=s['deg_status']
        sc(row,9,DEG_LABEL[ds],font=DEG_FONT[ds],fill=DEG_FILL[ds],aln=C)
        sc(row,10,s['last_dur'],font=F10,fill=bg,aln=C)
        sc(row,11,'',fill=SEP)

        def fill_str(start_col, s_num):
            v = s['strings'].get(s_num)
            if not v:
                for c in range(start_col, start_col+5): sc(row,c,'N/A',font=F9,fill=bg,aln=C)
                return
            st=v['status']
            sc(row,start_col,   STR_LABEL[st],         font=STR_FONT[st],fill=STR_FILL[st],aln=C)
            sc(row,start_col+1, f"{v['stopper_pct']}%",font=F10,fill=bg,aln=C)
            sc(row,start_col+2, f"{v['avg_cap']}%",    font=F10,fill=bg,aln=C)
            sc(row,start_col+3, f"{v['avg_volt']}V",   font=F10,fill=bg,aln=C)
            sc(row,start_col+4, v['note'],              font=F9, fill=bg,aln=L)

        fill_str(12,1)
        sc(row,17,'',fill=SEP)
        fill_str(18,2)

    ws.freeze_panes='A3'
    wb.save(out_path)
    print(f"\n Excel → {out_path}")

def collect_csv_files(paths):

    import glob as _glob
    files = []
    for p in paths:
        if os.path.isdir(p):
            found = sorted(_glob.glob(os.path.join(p, '*.csv')))
            print(f"Folder: {p}  ({len(found)} CSV файл олдлоо)")
            files.extend(found)
        elif '*' in p or '?' in p:
            found = sorted(_glob.glob(p))
            print(f"Pattern: {p}  ({len(found)} файл)")
            files.extend(found)
        elif os.path.isfile(p):
            files.append(p)
        else:
            print(f" Олдсонгүй: {p}")
    return files


def main():
    usage = """
Ашиглах (Usage):
  python battery_analysis.py <csv_файл1> [csv_файл2 ...]   — тодорхой файлууд
  python battery_analysis.py /path/to/csv_folder/          — бүх CSV бүхий folder
  python battery_analysis.py /path/*.csv                   — glob pattern
  python battery_analysis.py site1.csv site2.csv /folder/  — холимог

Жишээ (93 site):
  python battery_analysis.py C:/data/sites/
  python battery_analysis.py /home/user/discharge_data/*.csv
"""
    if len(sys.argv) < 2:
        print(usage); sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    csv_files  = collect_csv_files(sys.argv[1:])

    if not csv_files:
        print("CSV файл олдсонгүй."); sys.exit(1)

    print(f"\n{'='*55}")
    print(f"Нийт {len(csv_files)} CSV файл анализлана...")
    print(f"{'='*55}")

    all_sites = []
    for i, fp in enumerate(csv_files, 1):
        print(f"[{i:3d}/{len(csv_files)}]", end='')
        r = analyze_file(fp)
        if r:
            all_sites.append(r)

    if not all_sites:
        print("Анализлах өгөгдөл байхгүй."); sys.exit(1)

    out = os.path.join(script_dir, "battery_health_report.xlsx")
    build_excel(all_sites, out)
    print(f"\nНийт {len(all_sites)} site анализлав.")


if __name__=='__main__': main()