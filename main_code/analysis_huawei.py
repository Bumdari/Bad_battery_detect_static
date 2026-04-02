import pandas as pd
import numpy as np
import os
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

data_folder = 'data_huawei'
OUTPUT_FILE  = 'output_analysis_huawei.xlsx'

MIN_DURATION    = 13
BASELINE_TOP_N  = 10
SLOPE_DAYS      = 30
MIN_CYCLES_PRED = 5
MAX_PRED_DAYS   = 365
PRED_WINDOW     = 20

LEVEL_CRITICAL = 30
LEVEL_AVERAGE  = 60
LEVEL_GOOD     = 120
LEVEL_VERYGOOD = 180

INVALID_REASONS = ['Low temperature', 'Fault alarm']
VALID_REASONS   = ['Low voltage', 'Low capacity', 'Exceeded time limit']


def duration_level(dur):
    if dur < LEVEL_CRITICAL:
        return 'CRITICAL'
    elif dur <= LEVEL_AVERAGE:
        return 'AVERAGE'
    elif dur <= LEVEL_GOOD:
        return 'GOOD'
    else:
        return 'VERY GOOD'

def is_valid_cycle(row):
    if row['stop_reason'] in INVALID_REASONS:
        return False
    if row['test_type'] != 'Scheduled test':
        return False
    if row['duration_min'] < MIN_DURATION and row['stop_reason'] not in ('Low voltage', 'Exceeded time limit'):
        return False
    return True

def weighted_regression(df_window):
    if len(df_window) < MIN_CYCLES_PRED:
        return None, None, None

    x = np.arange(len(df_window), dtype=float)
    y = df_window['duration_min'].values.astype(float)
    w = x + 1

    slope, intercept = np.polyfit(x, y, 1, w=w)

    dates = df_window['start_time'].values
    if len(dates) >= 2:
        total_days = (pd.Timestamp(dates[-1]) - pd.Timestamp(dates[0])).days
        avg_days_per_cycle = total_days / (len(dates) - 1) if total_days > 0 else 2
    else:
        avg_days_per_cycle = 2

    return round(slope, 4), round(intercept, 4), round(avg_days_per_cycle, 2)

def predict_days_to_threshold(current_idx, slope, intercept, threshold, avg_days_per_cycle, today):
    if slope is None or slope >= 0:
        return 'Сайжирч байна'

    current_pred = slope * current_idx + intercept
    if current_pred <= threshold:
        return 'Доошилж эхэлсэн'

    cycles_needed = (current_pred - threshold) / abs(slope)
    days_needed = cycles_needed * avg_days_per_cycle

    if days_needed > MAX_PRED_DAYS:
        return '1 жилд буурахгүй'
    if days_needed <= 0:
        return 'Доошилж эхэлсэн'

    pred_date = today + timedelta(days=int(days_needed))
    return pred_date.strftime('%Y-%m-%d')

def get_trend(slope):
    if slope is None:
        return 'Insufficient data'
    if slope > 0.3:
        return 'IMPROVING'
    elif slope < -0.3:
        return 'DECLINING'
    else:
        return 'STABLE'


all_files = [f for f in os.listdir(data_folder) if f.lower().endswith('.csv')]
print(f"{len(all_files)} ширхэг csv")

results = []
today = pd.Timestamp.today().normalize()

for filename in all_files:
    filepath = os.path.join(data_folder, filename)
    try:
        df = pd.read_csv(filepath)
        df.columns = df.columns.str.strip().str.lstrip('\ufeff')
        df['start_time'] = pd.to_datetime(df['start_time'])
        df['end_time']   = pd.to_datetime(df['end_time'])
        df = df.sort_values('start_time').reset_index(drop=True)
        df['duration_min'] = (df['end_time'] - df['start_time']).dt.total_seconds() / 60
        df['duration_min'] = df['duration_min'].round(2)

        df_all = df[df['stop_reason'].isin(VALID_REASONS + INVALID_REASONS)].reset_index(drop=True)
        if len(df_all) < 1:
            print(f" → {filename}: хоосон, алгасав")
            continue

        battery_change_indices = [0]
        durations_raw = df_all['duration_min']
        for i in range(1, len(df_all)):
            if i >= 5 and i + 5 <= len(df_all):
                pre_vals    = durations_raw.iloc[i-5:i]
                post_vals   = durations_raw.iloc[i:i+5]
                pre_nonzero = pre_vals[pre_vals > 0]
                post_med    = post_vals.median()
                pre_med     = pre_nonzero.median() if len(pre_nonzero) > 0 else 0
                if pre_med > 0 and post_med / pre_med >= 1.5 and post_med >= 60:
                    battery_change_indices.append(i)
                elif len(pre_nonzero) == 0 and post_med >= 60:
                    battery_change_indices.append(i)
        last_change_time = df_all.loc[battery_change_indices[-1], 'start_time']

        df = df_all[df_all.apply(is_valid_cycle, axis=1)].reset_index(drop=True)
        if len(df) < 1:
            print(f" → {filename}: шүүлтийн дараа хоосон")
            continue

        battery_change_date = last_change_time
        recent_df = df[df['start_time'] >= last_change_time].copy().reset_index(drop=True)

        top_10   = recent_df['duration_min'].sort_values(ascending=False).head(BASELINE_TOP_N)
        baseline = top_10.median()

        cutoff_date = df['start_time'].max() - pd.Timedelta(days=SLOPE_DAYS)
        last_30_df  = df[df['start_time'] >= cutoff_date]
        slope_30    = last_30_df['duration_min'].median() if len(last_30_df) > 0 else recent_df['duration_min'].median()

        last_row_valid  = df.iloc[-1]
        last_row_all    = df_all.iloc[-1]
        last_duration   = last_row_valid['duration_min']
        last_stop_cause = last_row_all['stop_reason']
        last_avg_current  = last_row_all['avg_current']
        last_start_date = last_row_valid['start_time'].strftime('%Y-%m-%d')
        ip              = df_all['ip'].iloc[0]
        location        = df_all['location'].iloc[0]

        drop_percent = round(((baseline - slope_30) / baseline * 100), 2) if baseline > 0 else 0.0
        if drop_percent < 25:
            drop_level = 'NORMAL'
        elif drop_percent <= 55:
            drop_level = 'STRONG'
        else:
            drop_level = 'CRITICAL'

        if last_duration >= 59:
            status = 'stable'
        elif last_duration < 30:
            status = 'critical'
        else:
            status = 'degrading'

        min_idx      = recent_df['duration_min'].idxmin()
        min_duration = recent_df.loc[min_idx, 'duration_min']
        min_stop     = recent_df.loc[min_idx, 'stop_reason']
        min_date     = recent_df.loc[min_idx, 'start_time'].strftime('%Y-%m-%d')

        current_level = duration_level(last_duration)

        pred_window_df = recent_df.tail(PRED_WINDOW).reset_index(drop=True)
        wlr_slope, wlr_intercept, avg_days_cycle = weighted_regression(pred_window_df)
        trend       = get_trend(wlr_slope)
        current_idx = len(pred_window_df)

        has_enough = wlr_slope is not None and len(pred_window_df) >= MIN_CYCLES_PRED

        if current_level == 'VERY GOOD' and has_enough:
            pred_to_good     = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_GOOD,     avg_days_cycle, today)
            pred_to_average  = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_AVERAGE,  avg_days_cycle, today)
            pred_to_critical = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_CRITICAL, avg_days_cycle, today)
        elif current_level == 'GOOD' and has_enough:
            pred_to_good     = 'Одоогийн түвшин'
            pred_to_average  = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_AVERAGE,  avg_days_cycle, today)
            pred_to_critical = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_CRITICAL, avg_days_cycle, today)
        elif current_level == 'AVERAGE' and has_enough:
            pred_to_good     = 'Муу'
            pred_to_average  = 'Одоогийн түвшин'
            pred_to_critical = predict_days_to_threshold(current_idx, wlr_slope, wlr_intercept, LEVEL_CRITICAL, avg_days_cycle, today)
        elif current_level == 'CRITICAL':
            pred_to_good     = 'Хурцадмал'
            pred_to_average  = 'Хурцадмал'
            pred_to_critical = 'Хурцадмал'
        else:
            pred_to_good     = 'Өгөгдөл хүрэлцэхгүй'
            pred_to_average  = 'Өгөгдөл хүрэлцэхгүй'
            pred_to_critical = 'Өгөгдөл хүрэлцэхгүй'

        NEAR_CRITICAL_WINDOW = 10
        NEAR_CRITICAL_COUNT  = 2
        recent_10    = recent_df.tail(NEAR_CRITICAL_WINDOW)
        critical_hits = (recent_10['duration_min'] < LEVEL_CRITICAL).sum()
        near_critical = critical_hits >= NEAR_CRITICAL_COUNT

        if near_critical:
            last_pred = recent_10['duration_min'].mean()
            if wlr_slope is not None and wlr_slope < 0:
                cycles_needed = (last_pred - LEVEL_CRITICAL) / abs(wlr_slope)
                days_needed   = cycles_needed * avg_days_cycle
                if 0 < days_needed <= MAX_PRED_DAYS:
                    near_critical_date = (today + timedelta(days=int(days_needed))).strftime('%Y-%m-%d')
                else:
                    near_critical_date = 'Тодорхойгүй'
            else:
                near_critical_date = (today + timedelta(days=int(avg_days_cycle * 3))).strftime('%Y-%m-%d')
        else:
            near_critical_date = None

        trend_arrow = '↘' if trend == 'DECLINING' else ('↗' if trend == 'IMPROVING' else '→')

        if current_level == 'VERY GOOD':
            next_level = 'GOOD'
            next_date  = pred_to_good
        elif current_level == 'GOOD':
            next_level = 'AVERAGE'
            next_date  = pred_to_average
        elif current_level == 'AVERAGE':
            next_level = 'CRITICAL'
            next_date  = pred_to_critical
        else:
            next_level = None
            next_date  = None

        if current_level == 'CRITICAL':
            forecast_str = 'CRITICAL ↘ Хурцадмал'
        elif near_critical:
            forecast_str = f"{current_level} ↘ CRITICAL орчимд  ({near_critical_date})"
        elif next_level and next_date and next_date not in (
                'Сайжирч байна', 'Доошилж эхэлсэн', '1 жилд буурахгүй',
                'Муу', 'Одоогийн түвшин', 'Хурцадмал', 'Өгөгдөл хүрэлцэхгүй'):
            forecast_str = f"{current_level} {trend_arrow} {next_level} ({next_date})"
        elif trend in ('IMPROVING', 'STABLE'):
            forecast_str = f"{current_level} {trend_arrow} Хэвийн"
        else:
            forecast_str = f"{current_level} {trend_arrow} {next_date or '—'}"

        results.append({
            'ip address'               : ip,
            'site location'            : location,
            'battery change date'      : battery_change_date.strftime('%Y-%m-%d'),
            'baseline'                 : round(baseline, 2),
            'slope(30d median)'        : round(slope_30, 2),
            'drop percent'             : drop_percent,
            'last start date'          : last_start_date,
            'last duration_min'        : round(last_duration, 1),
            'last stop cause'          : last_stop_cause,
            'last avg_current'         : last_avg_current,
            'critical degrading stable': status,
            'status & forecast'        : forecast_str,
            '_drop_level'              : drop_level,
        })

        print(f" ✓ {filename} → {ip} | {current_level} | {trend} | drop: {drop_percent}%")

    except Exception as e:
        print(f" ✗ {filename} алдаа: {e}")

if not results:
    print("Ямар ч файл боловсруулж чадсангүй!")
else:
    output_df = pd.DataFrame(results)

    status_order = {'critical': 0, 'degrading': 1, 'stable': 2}
    output_df['_sort'] = output_df['critical degrading stable'].map(status_order)
    output_df = output_df.sort_values('_sort').drop(columns='_sort').reset_index(drop=True)

    drop_level_map = output_df['_drop_level'].tolist()
    output_df = output_df.drop(columns=['_drop_level'])

    output_df.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    col_map = {}
    for c in range(1, ws.max_column + 1):
        col_map[ws.cell(1, c).value] = c

    red    = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
    orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    yellow = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    green  = PatternFill(start_color='00CC66', end_color='00CC66', fill_type='solid')

    for i, row in enumerate(range(2, ws.max_row + 1)):

        if 'critical degrading stable' in col_map:
            val = ws.cell(row, col_map['critical degrading stable']).value
            if val == 'critical':    ws.cell(row, col_map['critical degrading stable']).fill = red
            elif val == 'degrading': ws.cell(row, col_map['critical degrading stable']).fill = orange
            elif val == 'stable':    ws.cell(row, col_map['critical degrading stable']).fill = green

        if 'drop percent' in col_map:
            dl = drop_level_map[i]
            if dl == 'CRITICAL': ws.cell(row, col_map['drop percent']).fill = red
            elif dl == 'STRONG': ws.cell(row, col_map['drop percent']).fill = orange
            else:                ws.cell(row, col_map['drop percent']).fill = green

        if 'status & forecast' in col_map:
            val = ws.cell(row, col_map['status & forecast']).value or ''
            if val.startswith('CRITICAL') or 'CRITICAL орчимд' in val:
                ws.cell(row, col_map['status & forecast']).fill = red
            elif val.startswith('AVERAGE'):   ws.cell(row, col_map['status & forecast']).fill = orange
            elif val.startswith('GOOD'):      ws.cell(row, col_map['status & forecast']).fill = yellow
            elif val.startswith('VERY GOOD'): ws.cell(row, col_map['status & forecast']).fill = green

        if 'last stop cause' in col_map:
            val = ws.cell(row, col_map['last stop cause']).value or ''
            if val in INVALID_REASONS:
                ws.cell(row, col_map['last stop cause']).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = header_fill
        ws.cell(1, c).font = header_font
        ws.cell(1, c).alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = 24

    ws.row_dimensions[1].height = 35
    wb.save(OUTPUT_FILE)

    print(f"\n{'='*55}")
    print(f"ДУУСЛАА: {len(results)} IP → {OUTPUT_FILE}")
